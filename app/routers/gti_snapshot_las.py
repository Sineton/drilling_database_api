"""Dedicated router for mapping/importing LAS data into gti_snapshots."""
from typing import Any, Dict, List, Optional, Tuple
import os
from datetime import datetime, timedelta
from pathlib import Path

import lasio
from openpyxl import load_workbook

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import desc

from ..database import get_db
from ..models import GtiLog, LogChannel, Well, Wellbore
from ..models.gti_snapshot import GtiSnapshot
from ..services.las_parser import LASParserService


router = APIRouter(prefix="/gti-snapshot-las", tags=["GTI Snapshot LAS"])


WORKBOOK_CANDIDATE_PATHS = [
    Path(__file__).resolve().parents[3] / "las_parameters.xlsx",
    Path(__file__).resolve().parents[3] / "las_parameters.xlxs",
    Path(__file__).resolve().parents[3] / "Таблица_соответствия_обозначений_параметров.xlsx",
]

# Explicit LAS mnemonic aliases -> workbook abbreviation.
# This uses the example file conventions and improves matching quality.
MNEMONIC_TO_ABBREVIATION = {
    "Zab": "DMEA",
    "Gl.dol": "DBTM",
    "W": "WOBA",
    "Hkr": "BPOS",
    "M": "TQA",
    "W kr": "HKLA",
    "P vkh": "SPPA",
    "N rot": "RPMA",
    "V sum": "TVT",
    "Q vkh": "MFIA",
    "Q vyikh": "MFOA",
    "G vkh": "MDIA",
    "G vyikh": "MDOA",
    "G sum": "GASA",
}

# Mnemonics that must stay unmapped into snapshot base columns.
# They should be routed to params_extra on import stage.
FORCE_UNMAPPED_MNEMONICS = {
    "m kl",
}


class SnapshotMappingRequest(BaseModel):
    """Request for LAS->gti_snapshot mapping preview."""
    file_path: str = Field(..., description="Absolute path to LAS file")
    well_number: Optional[str] = Field(default=None, description="Optional explicit well number")
    create_well: bool = Field(default=False, description="Create well when not found")
    channel_mapping: Optional[Dict[str, str]] = Field(
        default=None,
        description="Optional mapping override: LAS mnemonic -> gti_snapshot field",
    )


class SnapshotImportRequest(BaseModel):
    """Request for direct LAS import into gti_snapshots."""
    file_path: str = Field(..., description="Absolute path to LAS file")
    log_id: Optional[int] = Field(default=None, description="Prepared GTI log ID from /mapping")
    well_number: Optional[str] = Field(default=None, description="Optional explicit well number")
    create_well: bool = Field(default=False, description="Create well when not found")
    channel_mapping: Optional[Dict[str, str]] = Field(
        default=None,
        description="Optional mapping override: LAS mnemonic -> gti_snapshot field",
    )
    batch_size: int = Field(default=50000, description="Batch insert size")


def _normalize_mnemonic(value: str) -> str:
    # LAS mnemonic variants may differ by spacing/dots/case (e.g. "P vkh.", "P vkh").
    return (
        (value or "")
        .strip()
        .lower()
        .replace(".", "")
        .replace("_", " ")
    )


def _normalize_code(value: str) -> str:
    if value is None:
        return ""
    # Handle Cyrillic "С" in "С1-С5" and normalize symbols.
    text = (
        str(value)
        .strip()
        .upper()
        .replace("С", "C")
        .replace("-", "")
        .replace("_", "")
        .replace(".", "")
        .replace(" ", "")
    )
    return text


def _to_json_scalar(value: Any) -> Any:
    """Convert numpy/pandas scalars to plain Python values."""
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def _get_workbook_path() -> Optional[Path]:
    for path in WORKBOOK_CANDIDATE_PATHS:
        if path.exists():
            return path
    return None


def _load_abbreviation_to_snapshot_field() -> Tuple[Dict[str, str], Optional[str]]:
    workbook_path = _get_workbook_path()
    if workbook_path is None:
        return {}, None

    valid_snapshot_fields = set(GtiSnapshot.__table__.columns.keys())
    normalized_fields = {_normalize_code(field): field for field in valid_snapshot_fields}

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    result: Dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        abbr = row[2]
        if not abbr:
            continue
        normalized_abbr = _normalize_code(abbr)
        target = normalized_fields.get(normalized_abbr)
        if target:
            result[normalized_abbr] = target

    return result, str(workbook_path)


def _build_effective_mapping(
    curves: List[Dict[str, Any]],
    mapping_override: Optional[Dict[str, str]] = None,
) -> Tuple[Dict[str, str], Dict[str, Any]]:
    base_mapping = {**LASParserService.DEFAULT_CHANNEL_MAPPING}
    normalized_lookup = {_normalize_mnemonic(k): v for k, v in base_mapping.items()}
    normalized_aliases = {
        _normalize_mnemonic(k): _normalize_code(v)
        for k, v in MNEMONIC_TO_ABBREVIATION.items()
    }
    abbr_to_field, workbook_path = _load_abbreviation_to_snapshot_field()
    valid_snapshot_fields = set(GtiSnapshot.__table__.columns.keys())
    override_mapping = mapping_override or {}

    effective: Dict[str, str] = {}
    source_stats = {"override": 0, "workbook": 0, "default": 0}

    for curve in curves:
        mnemonic = curve.get("mnemonic") or ""
        normalized_mnemonic = _normalize_mnemonic(mnemonic)

        # Explicit guardrail: keep selected channels only in params_extra.
        if normalized_mnemonic in FORCE_UNMAPPED_MNEMONICS:
            continue

        mapped_field = None

        # 1) User-provided override has highest priority.
        if mnemonic in override_mapping and override_mapping[mnemonic] in valid_snapshot_fields:
            mapped_field = override_mapping[mnemonic]
            source_stats["override"] += 1
        elif normalized_mnemonic in {
            _normalize_mnemonic(k): v for k, v in override_mapping.items()
        }:
            candidate = {
                _normalize_mnemonic(k): v for k, v in override_mapping.items()
            }[normalized_mnemonic]
            if candidate in valid_snapshot_fields:
                mapped_field = candidate
                source_stats["override"] += 1

        # 2) Workbook-based mapping via mnemonic alias.
        if not mapped_field:
            abbr = normalized_aliases.get(normalized_mnemonic)
            if abbr:
                candidate = abbr_to_field.get(abbr)
                if candidate in valid_snapshot_fields:
                    mapped_field = candidate
                    source_stats["workbook"] += 1

        # 3) Fallback to legacy/default mapping from parser service.
        if not mapped_field:
            direct = base_mapping.get(mnemonic)
            normalized = normalized_lookup.get(normalized_mnemonic)
            candidate = direct or normalized
            if candidate in valid_snapshot_fields:
                mapped_field = candidate
                source_stats["default"] += 1

        if mapped_field and mapped_field in valid_snapshot_fields:
            effective[mnemonic] = mapped_field

    return effective, {
        "workbook_path": workbook_path,
        "source_stats": source_stats,
    }


def _parse_las_header_fast(file_path: str) -> Dict[str, Any]:
    """Read LAS header/curves only (without ASCII data section)."""
    las = lasio.read(file_path, ignore_data=True)

    well_info: Dict[str, Any] = {}
    for item in las.well:
        well_info[item.mnemonic] = _to_json_scalar(item.value)

    curves: List[Dict[str, Any]] = []
    for curve in las.curves:
        curves.append(
            {
                "mnemonic": curve.mnemonic,
                "unit": curve.unit or "",
                "description": curve.descr or "",
            }
        )

    return {
        "las_version": las.version[0].value if las.version else "2.0",
        "well_info": well_info,
        "curves": curves,
    }


def _resolve_well_number(parsed_well_info: Dict[str, Any], file_path: str, explicit_well_number: Optional[str]) -> str:
    if explicit_well_number:
        return str(explicit_well_number).strip()
    if parsed_well_info.get("WELL"):
        return str(parsed_well_info.get("WELL")).strip()
    return Path(file_path).parent.name


def _extract_sampling_rate_sec(step_value: Any) -> float:
    if step_value is None:
        return 1.0
    try:
        step_str = str(step_value).lower().replace("sec", "").strip()
        return float(step_str) if step_str else 1.0
    except Exception:
        return 1.0


def _prepare_log_context(
    db: Session,
    file_path: str,
    parsed: Dict[str, Any],
    well_number: Optional[str],
    create_well: bool,
) -> GtiLog:
    service = LASParserService(db)
    well_info = parsed.get("well_info", {})
    resolved_well_number = _resolve_well_number(well_info, file_path, well_number)

    well = service._find_well(resolved_well_number)
    if not well:
        if not create_well:
            raise HTTPException(status_code=400, detail=f"Well not found: {resolved_well_number}")
        well = Well(
            well_number=resolved_well_number,
            field=str(well_info.get("FLD") or ""),
            project_code="imported",
            company=str(well_info.get("COMP") or "unknown"),
            metadata_={
                "kust": str(well_info.get("KUST") or ""),
                "srvc": str(well_info.get("SRVC") or ""),
            },
        )
        db.add(well)
        db.commit()
        db.refresh(well)

    wellbore = db.query(Wellbore).filter(
        Wellbore.well_id == well.well_id,
        Wellbore.wellbore_number == "main",
    ).first()
    if not wellbore:
        wellbore = Wellbore(well_id=well.well_id, wellbore_number="main")
        db.add(wellbore)
        db.commit()
        db.refresh(wellbore)

    start_time = service._parse_las_datetime(well_info.get("STRT")) or datetime.utcnow()
    end_time = service._parse_las_datetime(well_info.get("STOP")) or (start_time + timedelta(seconds=1))
    if end_time <= start_time:
        end_time = start_time + timedelta(seconds=1)

    sampling_rate = _extract_sampling_rate_sec(well_info.get("STEP"))

    existing_log = db.query(GtiLog).filter(
        GtiLog.wellbore_id == wellbore.wellbore_id,
        GtiLog.source_file_path == file_path,
        GtiLog.file_format == "las",
        GtiLog.quality_status == "prepared",
    ).order_by(desc(GtiLog.log_id)).first()

    if existing_log:
        existing_log.start_time = start_time
        existing_log.end_time = end_time
        existing_log.sampling_rate_sec = sampling_rate
        existing_log.quality_status = "prepared"
        existing_log.source_file_path = file_path
        existing_log.file_format = "las"
        db.commit()
        db.refresh(existing_log)
        return existing_log

    prepared_log = GtiLog(
        wellbore_id=wellbore.wellbore_id,
        start_time=start_time,
        end_time=end_time,
        sampling_rate_sec=sampling_rate,
        total_records=0,
        quality_status="prepared",
        source_file_path=file_path,
        file_format="las",
    )
    db.add(prepared_log)
    db.commit()
    db.refresh(prepared_log)
    return prepared_log


def _upsert_log_channels(
    db: Session,
    log_id: int,
    curves: List[Dict[str, Any]],
    effective_mapping: Dict[str, str],
) -> Dict[str, int]:
    existing_rows = db.query(LogChannel).filter(LogChannel.log_id == log_id).all()
    existing_by_mnemonic = {row.las_mnemonic: row for row in existing_rows}
    seen_mnemonics = set()
    inserted = 0
    updated = 0

    for curve in curves:
        mnemonic = curve.get("mnemonic") or ""
        if not mnemonic:
            continue
        seen_mnemonics.add(mnemonic)
        mapped_to = effective_mapping.get(mnemonic)
        status = "mapped" if mapped_to else "skipped"

        current = existing_by_mnemonic.get(mnemonic)
        if current:
            current.las_unit = curve.get("unit")
            current.las_description = curve.get("description")
            current.db_column_name = mapped_to
            current.import_status = status
            current.import_notes = None if mapped_to else "No effective mapping found"
            updated += 1
        else:
            db.add(
                LogChannel(
                    log_id=log_id,
                    las_mnemonic=mnemonic,
                    las_unit=curve.get("unit"),
                    las_description=curve.get("description"),
                    db_column_name=mapped_to,
                    import_status=status,
                    import_notes=None if mapped_to else "No effective mapping found",
                )
            )
            inserted += 1

    for row in existing_rows:
        if row.las_mnemonic not in seen_mnemonics:
            db.delete(row)

    db.commit()
    return {"inserted": inserted, "updated": updated}


@router.post("/mapping")
def preview_snapshot_mapping(
    request: SnapshotMappingRequest,
    db: Session = Depends(get_db),
):
    """Preview mapping and persist it into log_channels for reuse on import."""
    if not os.path.exists(request.file_path):
        raise HTTPException(status_code=404, detail=f"File not found: {request.file_path}")

    parsed = _parse_las_header_fast(request.file_path)
    prepared_log = _prepare_log_context(
        db=db,
        file_path=request.file_path,
        parsed=parsed,
        well_number=request.well_number,
        create_well=request.create_well,
    )
    effective_mapping, mapping_meta = _build_effective_mapping(parsed["curves"], request.channel_mapping)
    upsert_stats = _upsert_log_channels(db, prepared_log.log_id, parsed["curves"], effective_mapping)

    mapped_curves = []
    unmapped_curves = []
    for curve in parsed["curves"]:
        mnemonic = curve.get("mnemonic")
        item = {
            "mnemonic": mnemonic,
            "unit": curve.get("unit"),
            "description": curve.get("description"),
            "mapped_to": effective_mapping.get(mnemonic),
        }
        if item["mapped_to"]:
            mapped_curves.append(item)
        else:
            unmapped_curves.append(item)

    well_info = parsed.get("well_info", {})
    metadata_mapping = {
        "WELL": {"value": well_info.get("WELL"), "used_as": "well_number"},
        "STRT": {"value": well_info.get("STRT"), "used_as": "gti_logs.start_time"},
        "STOP": {"value": well_info.get("STOP"), "used_as": "gti_logs.end_time"},
        "STEP": {"value": well_info.get("STEP"), "used_as": "gti_logs.sampling_rate_sec"},
    }

    return {
        "success": True,
        "log_id": prepared_log.log_id,
        "file_path": request.file_path,
        "records_total": None,
        "parse_mode": "header_only_fast",
        "mapping_source": mapping_meta,
        "mapping_persisted": {
            "table": "log_channels",
            "upsert": upsert_stats,
        },
        "metadata_mapping": metadata_mapping,
        "curve_mapping": {
            "mapped_count": len(mapped_curves),
            "unmapped_count": len(unmapped_curves),
            "mapped": mapped_curves,
            "unmapped": unmapped_curves,
            "effective_mapping": effective_mapping,
        },
    }


@router.post("/import")
def import_las_to_gti_snapshot(
    request: SnapshotImportRequest,
    db: Session = Depends(get_db),
):
    """Import LAS file into gti_snapshots using mapping from log_channels when available."""
    if not os.path.exists(request.file_path):
        raise HTTPException(status_code=404, detail=f"File not found: {request.file_path}")

    service = LASParserService(db)
    mapping_meta: Dict[str, Any] = {}
    prepared_log: Optional[GtiLog] = None

    if request.log_id is not None:
        prepared_log = db.query(GtiLog).filter(GtiLog.log_id == request.log_id).first()
        if not prepared_log:
            raise HTTPException(status_code=404, detail=f"Prepared log not found: {request.log_id}")
    else:
        prepared_log = db.query(GtiLog).filter(
            GtiLog.source_file_path == request.file_path,
            GtiLog.file_format == "las",
            GtiLog.quality_status == "prepared",
        ).order_by(desc(GtiLog.log_id)).first()

    effective_mapping: Dict[str, str] = {}
    unmapped_columns_for_extra: List[str] = []
    if prepared_log:
        mapped_rows = db.query(LogChannel).filter(
            LogChannel.log_id == prepared_log.log_id,
            LogChannel.db_column_name.isnot(None),
        ).all()
        effective_mapping = {row.las_mnemonic: row.db_column_name for row in mapped_rows if row.db_column_name}

        unmapped_rows = db.query(LogChannel).filter(
            LogChannel.log_id == prepared_log.log_id,
            LogChannel.db_column_name.is_(None),
        ).all()
        unmapped_columns_for_extra = [row.las_mnemonic for row in unmapped_rows if row.las_mnemonic]
        mapping_meta = {
            "source": "log_channels",
            "log_id": prepared_log.log_id,
            "mapped_channels": len(effective_mapping),
            "unmapped_channels_to_params_extra": len(unmapped_columns_for_extra),
        }

    if request.channel_mapping:
        effective_mapping.update(request.channel_mapping)
        mapping_meta["override_applied"] = True

    if not effective_mapping:
        parsed = _parse_las_header_fast(request.file_path)
        effective_mapping, build_meta = _build_effective_mapping(parsed["curves"], request.channel_mapping)
        mapping_meta = {"source": "calculated_on_import", **build_meta}

    result = service.import_las(
        file_path=request.file_path,
        well_number=request.well_number,
        create_well=request.create_well,
        channel_mapping=effective_mapping,
        extra_columns=unmapped_columns_for_extra,
        batch_size=request.batch_size,
        existing_log_id=prepared_log.log_id if prepared_log else None,
    )

    if prepared_log:
        db.query(LogChannel).filter(
            LogChannel.log_id == prepared_log.log_id,
            LogChannel.db_column_name.isnot(None),
        ).update(
            {
                LogChannel.import_status: "imported",
                LogChannel.import_notes: "Imported via /gti-snapshot-las/import",
            },
            synchronize_session=False,
        )
        db.query(LogChannel).filter(
            LogChannel.log_id == prepared_log.log_id,
            LogChannel.db_column_name.is_(None),
        ).update(
            {
                LogChannel.import_status: "imported",
                LogChannel.import_notes: "Stored in gti_snapshots.params_extra",
            },
            synchronize_session=False,
        )
        db.commit()

    return {
        "success": True,
        "mapping_source": mapping_meta,
        "mapping_used": effective_mapping,
        "log_id_used": result.get("log_id"),
        "import_result": result,
    }
