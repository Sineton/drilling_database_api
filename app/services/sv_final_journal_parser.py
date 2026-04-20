"""
Импорт мультилистового «финального» журнала (final.xlsx): листы
Баланс, График, Детализация, Инциденты.

Заголовок «Суточный отчёт» на листе «Отчёт» — шаблон; период задаётся
данными на листах Баланс/Детализация/Инциденты (и при необходимости
построчными датами в «Примечание» на «Отчёте» — см. отдельные задачи).
"""
from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from ..models import (
    Well, Wellbore, File,
    SvDailyReport, SvDailyOperation, SvNpvBalance,
)
from .sv_journal_parser import (
    _float,
    _int,
    _text,
    _parse_date,
    _parse_time,
    _classify_operation,
    _detect_anomalies,
    _extract_params,
    _extract_depth_range,
    _parse_duration_minutes,
)

logger = logging.getLogger(__name__)

SHEET_BALANCE = "Баланс"
SHEET_GRAPH = "График "  # пробел в конце — как в файле
SHEET_DETAIL = "Детализация"
SHEET_INCIDENTS = "Инциденты"


def _cell(ws: Worksheet, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def _to_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return _parse_date(val)


def _time_or_datetime_to_time(val) -> Optional[time]:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    return _parse_time(val)


def _duration_cell_to_minutes(val) -> Optional[int]:
    if val is None:
        return None
    if isinstance(val, time):
        return val.hour * 60 + val.minute + (val.second // 60)
    if isinstance(val, datetime):
        return val.hour * 60 + val.minute
    s = _text(val)
    if not s:
        return None
    if re.match(r"^\d{1,2}:\d{2}", s):
        parts = s.replace(" ", "").split(":")
        try:
            h, m = int(parts[0]), int(parts[1]) if len(parts) > 1 else 0
            sec = int(parts[2]) if len(parts) > 2 else 0
            return h * 60 + m + sec // 60
        except (ValueError, IndexError):
            pass
    return _parse_duration_minutes(s)


def _hours_from_incident_cell(val) -> Optional[float]:
    """Колонка «Время НПВ» / календарные часы: float, time или строка."""
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return float(val)
    if isinstance(val, time):
        return val.hour + val.minute / 60.0 + val.second / 3600.0
    s = _text(str(val))
    if not s:
        return None
    f = _float(s.replace(",", "."))
    if f is not None:
        return f
    return None


def parse_well_from_detalization(ws: Worksheet) -> Dict[str, Any]:
    """Куст/скважина в шапке листа «Детализация» (строки 4–5)."""
    info: Dict[str, Any] = {}
    for row in range(1, 12):
        label = _text(_cell(ws, row, 2))
        val = _cell(ws, row, 3)
        if not label:
            continue
        ll = label.lower()
        if "куст" in ll:
            pad = _int(val)
            if pad is not None:
                info["pad_number"] = str(pad)
        if "скважин" in ll:
            wn = _text(val) or (str(val) if val is not None else None)
            if wn:
                info["well_number"] = wn.strip()
    return info


def parse_graph_notes(ws: Worksheet) -> Dict[date, str]:
    """Лист «График »: дата → примечание (колонка 5)."""
    notes: Dict[date, str] = {}
    header_row = None
    for row in range(1, min(ws.max_row + 1, 30)):
        v1 = _text(_cell(ws, row, 1))
        if v1 and v1.lower() == "дата":
            header_row = row
            break
    if not header_row:
        return notes
    for row in range(header_row + 1, ws.max_row + 1):
        d = _to_date(_cell(ws, row, 1))
        if not d:
            continue
        note = _text(_cell(ws, row, 5))
        if note:
            if d in notes:
                notes[d] = f"{notes[d]} | {note}"
            else:
                notes[d] = note
    return notes


def parse_balance_daily_rows(ws: Worksheet) -> List[Dict[str, Any]]:
    """
    Лист «Баланс»: строки с датой в колонке 1, забой/проходка, итого часов бурения (кол. 9).
    Пропускает строки-итоги без datetime (например «Всего под ОК-426»).
    """
    rows: List[Dict[str, Any]] = []
    for row in range(4, ws.max_row + 1):
        raw = _cell(ws, row, 1)
        d = _to_date(raw)
        if not d:
            continue
        depth = _float(_cell(ws, row, 2))
        penetration = _float(_cell(ws, row, 3))
        drilling_total = _float(_cell(ws, row, 9))
        rows.append({
            "row": row,
            "report_date": d,
            "current_depth_m": depth,
            "penetration_m": penetration,
            "drilling_time_h": drilling_total,
        })
    return rows


def iter_detalization_operations(ws: Worksheet) -> List[Dict[str, Any]]:
    """
    Разбор операций на «Детализация»: блоки с заголовком
    № п/п | Дата | От | До | Часов | Операция.
    """
    ops: List[Dict[str, Any]] = []
    in_table = False
    last_op_date: Optional[date] = None
    seq_counter = 0

    for row in range(1, ws.max_row + 1):
        h1 = _text(_cell(ws, row, 1))
        h2 = _text(_cell(ws, row, 2))

        if h2 and "дата" in h2.lower() and h1 and "п/п" in h1.lower():
            in_table = True
            seq_counter = 0
            continue

        if not in_table:
            continue

        c1 = _cell(ws, row, 1)
        c2 = _cell(ws, row, 2)
        c3 = _cell(ws, row, 3)
        c4 = _cell(ws, row, 4)
        c5 = _cell(ws, row, 5)
        c6 = _cell(ws, row, 6)

        d2 = _to_date(c2)
        if d2:
            last_op_date = d2

        desc = _text(c6)
        if desc and len(desc) > 5 and "отчет по строительству" in desc.lower():
            continue

        seq_val = _int(c1)
        if seq_val is not None and seq_val > 0 and last_op_date and desc:
            seq_counter = seq_val
            tf = _time_or_datetime_to_time(c3)
            tt = _time_or_datetime_to_time(c4)
            dur_txt = _text(c5)
            dur_min = _duration_cell_to_minutes(c5)

            if dur_min is None and tf and tt:
                t1 = datetime.combine(date.today(), tf)
                t2 = datetime.combine(date.today(), tt)
                if t2 < t1:
                    t2 += timedelta(days=1)
                dur_min = int((t2 - t1).total_seconds() / 60)

            full_desc = desc
            cat = _classify_operation(full_desc)
            flags, severity, is_npv, is_comp, is_rep = _detect_anomalies(full_desc)
            params = _extract_params(full_desc)
            dfm, dtm = _extract_depth_range(full_desc)

            ops.append({
                "report_date": last_op_date,
                "sequence_number": seq_val,
                "time_from": tf,
                "time_to": tt,
                "duration_text": dur_txt,
                "duration_minutes": dur_min,
                "description": full_desc,
                "operation_category": cat,
                "is_npv": is_npv,
                "is_complication": is_comp,
                "is_repair": is_rep,
                "extracted_params": params,
                "depth_from_m": dfm,
                "depth_to_m": dtm,
                "anomaly_flags": flags if flags else None,
                "anomaly_severity": severity,
                "source_row": row,
            })
        elif (
            last_op_date
            and desc
            and len(desc) > 3
            and (c1 is None or str(c1).strip() == "")
            and not d2
        ):
            # Продолжение описания предыдущей операции
            if ops:
                ops[-1]["description"] = f"{ops[-1]['description']} {desc}"

    return ops


def parse_incidents(ws: Worksheet) -> List[Dict[str, Any]]:
    """Лист «Инциденты»: строки с № в колонке 1."""
    items: List[Dict[str, Any]] = []
    start = None
    for row in range(1, min(ws.max_row + 1, 20)):
        v = _text(_cell(ws, row, 1))
        if v and "п/п" in v.lower():
            start = row + 1
            break
    if not start:
        return items

    for row in range(start, ws.max_row + 1):
        num = _int(_cell(ws, row, 1))
        if num is None or num < 1:
            continue

        kind = _text(_cell(ws, row, 2))
        typ = _text(_cell(ws, row, 3))
        descr = _text(_cell(ws, row, 4))
        measures = _text(_cell(ws, row, 5))
        t_start = _cell(ws, row, 6)
        t_end = _cell(ws, row, 7)
        npv_hours = _hours_from_incident_cell(_cell(ws, row, 9))

        inc_date = _to_date(t_start)
        if not inc_date:
            continue

        parts = [x for x in (descr, measures) if x]
        description = " — ".join(parts) if parts else (descr or typ or kind or "Инцидент")

        category = " / ".join(x for x in (kind, typ) if x) or "НПВ"

        items.append({
            "incident_date": inc_date,
            "description": description[:8000] if len(description) > 8000 else description,
            "duration_hours": npv_hours,
            "category": category[:500],
            "operation_type": typ,
            "root_cause": descr,
            "source_row": row,
        })
    return items


class SvFinalJournalParserService:
    """Импорт final.xlsx (мультилистовый журнал)."""

    def __init__(self, db: Session):
        self.db = db
        self.warnings: List[str] = []
        self.errors: List[str] = []

    def parse_preview(self, file_path: str) -> Dict[str, Any]:
        self.warnings = []
        self.errors = []
        wb = openpyxl.load_workbook(file_path, data_only=True)

        if SHEET_DETAIL not in wb.sheetnames:
            raise ValueError(f"В файле нет листа «{SHEET_DETAIL}»")

        ws_d = wb[SHEET_DETAIL]
        well_info = parse_well_from_detalization(ws_d)

        balance_rows: List[Dict[str, Any]] = []
        if SHEET_BALANCE in wb.sheetnames:
            balance_rows = parse_balance_daily_rows(wb[SHEET_BALANCE])

        graph_notes: Dict[date, str] = {}
        if SHEET_GRAPH in wb.sheetnames:
            graph_notes = parse_graph_notes(wb[SHEET_GRAPH])

        ops = iter_detalization_operations(ws_d)

        incidents: List[Dict[str, Any]] = []
        if SHEET_INCIDENTS in wb.sheetnames:
            incidents = parse_incidents(wb[SHEET_INCIDENTS])

        op_dates = [o["report_date"] for o in ops if o.get("report_date")]
        bal_dates = [r["report_date"] for r in balance_rows]
        all_dates = sorted(set(op_dates + bal_dates))
        date_range = None
        if all_dates:
            date_range = {"from": all_dates[0].isoformat(), "to": all_dates[-1].isoformat()}

        return {
            "success": True,
            "format": "final_multisheet",
            "file_info": {
                "filename": os.path.basename(file_path),
                "sheets": wb.sheetnames,
                "detail_rows": ws_d.max_row,
            },
            "well_info": well_info,
            "balance_rows": len(balance_rows),
            "graph_notes_days": len(graph_notes),
            "operations_count": len(ops),
            "incidents_count": len(incidents),
            "date_range": date_range,
            "warnings": self.warnings,
        }

    def import_journal(
        self,
        file_path: str,
        project_code: str = "pao-tatneft",
        well_number_override: Optional[str] = None,
        dry_run: bool = False,
    ) -> Dict[str, Any]:
        self.warnings = []
        self.errors = []

        wb = openpyxl.load_workbook(file_path, data_only=True)
        if SHEET_DETAIL not in wb.sheetnames:
            raise ValueError(f"В файле нет листа «{SHEET_DETAIL}»")

        ws_d = wb[SHEET_DETAIL]
        well_info = parse_well_from_detalization(ws_d)
        well_num = well_number_override or well_info.get("well_number")
        if not well_num:
            raise ValueError("Не удалось определить номер скважины. Укажите well_number.")

        balance_rows = (
            parse_balance_daily_rows(wb[SHEET_BALANCE]) if SHEET_BALANCE in wb.sheetnames else []
        )
        graph_notes = (
            parse_graph_notes(wb[SHEET_GRAPH]) if SHEET_GRAPH in wb.sheetnames else {}
        )
        operations = iter_detalization_operations(ws_d)
        incidents = (
            parse_incidents(wb[SHEET_INCIDENTS]) if SHEET_INCIDENTS in wb.sheetnames else []
        )

        summary = {
            "well_number": well_num,
            "daily_reports_created": 0,
            "operations_created": 0,
            "npv_records_created": 0,
            "warnings": self.warnings,
            "errors": self.errors,
            "daily_summaries": [],
        }

        if dry_run:
            summary["daily_reports_created"] = len({r["report_date"] for r in balance_rows}) or len(
                {o["report_date"] for o in operations if o.get("report_date")}
            )
            summary["operations_created"] = len(operations)
            summary["npv_records_created"] = len(incidents)
            return summary

        well = self._get_or_create_well(well_num, well_info, project_code)
        summary["well_id"] = well.well_id
        wellbore = self._get_or_create_wellbore(well)
        summary["wellbore_id"] = wellbore.wellbore_id

        if well_info.get("pad_number"):
            well.pad_number = well_info["pad_number"]
        well.has_supervision_log = True
        self.db.flush()

        file_record = self._register_file(file_path, well)

        # 1) Ежедневные отчёты из Баланс + примечание из График
        report_by_date: Dict[date, SvDailyReport] = {}
        for br in balance_rows:
            rd = br["report_date"]
            existing = self.db.query(SvDailyReport).filter(
                SvDailyReport.wellbore_id == wellbore.wellbore_id,
                SvDailyReport.report_date == rd,
            ).first()
            if existing:
                self.warnings.append(f"Отчёт за {rd} уже есть, обновляем поля из Баланса")
                rep = existing
            else:
                rep = SvDailyReport(
                    wellbore_id=wellbore.wellbore_id,
                    report_date=rd,
                    source_file_id=file_record.file_id if file_record else None,
                    source_row_start=br.get("row"),
                )
                self.db.add(rep)
                self.db.flush()
                summary["daily_reports_created"] += 1

            rep.current_depth_m = br.get("current_depth_m")
            rep.penetration_m = br.get("penetration_m")
            rep.drilling_time_h = br.get("drilling_time_h")
            note = graph_notes.get(rd)
            if note:
                if rep.drilling_comment:
                    rep.drilling_comment = f"{rep.drilling_comment}\n{note}"
                else:
                    rep.drilling_comment = note
            report_by_date[rd] = rep

        # Даты только из операций (если нет строки в Балансе)
        op_dates = sorted({o["report_date"] for o in operations if o.get("report_date")})
        for rd in op_dates:
            if rd in report_by_date:
                continue
            existing = self.db.query(SvDailyReport).filter(
                SvDailyReport.wellbore_id == wellbore.wellbore_id,
                SvDailyReport.report_date == rd,
            ).first()
            if existing:
                report_by_date[rd] = existing
                continue
            rep = SvDailyReport(
                wellbore_id=wellbore.wellbore_id,
                report_date=rd,
                source_file_id=file_record.file_id if file_record else None,
            )
            self.db.add(rep)
            self.db.flush()
            summary["daily_reports_created"] += 1
            note = graph_notes.get(rd)
            if note:
                rep.drilling_comment = note
            report_by_date[rd] = rep

        # 2) Операции
        for op in operations:
            rd = op.get("report_date")
            if not rd:
                self.warnings.append(f"Пропуск операции без даты (строка {op.get('source_row')})")
                continue
            rep = report_by_date.get(rd)
            if not rep:
                rep = self.db.query(SvDailyReport).filter(
                    SvDailyReport.wellbore_id == wellbore.wellbore_id,
                    SvDailyReport.report_date == rd,
                ).first()
                if not rep:
                    rep = SvDailyReport(
                        wellbore_id=wellbore.wellbore_id,
                        report_date=rd,
                        source_file_id=file_record.file_id if file_record else None,
                    )
                    self.db.add(rep)
                    self.db.flush()
                    summary["daily_reports_created"] += 1
                    report_by_date[rd] = rep
                else:
                    report_by_date[rd] = rep

            row_data = {k: v for k, v in op.items() if k != "report_date"}
            self.db.add(SvDailyOperation(report_id=rep.report_id, **row_data))
            summary["operations_created"] += 1

        # 3) НПВ
        for inc in incidents:
            self.db.add(
                SvNpvBalance(
                    wellbore_id=wellbore.wellbore_id,
                    source_file_id=file_record.file_id if file_record else None,
                    incident_date=inc["incident_date"],
                    description=inc["description"],
                    duration_hours=inc.get("duration_hours"),
                    category=inc["category"],
                    operation_type=inc.get("operation_type"),
                    root_cause=inc.get("root_cause"),
                    source_row=inc.get("source_row"),
                )
            )
            summary["npv_records_created"] += 1

        self.db.commit()

        if file_record:
            file_record.processing_status = "completed"
            self.db.commit()

        for rd, rep in sorted(report_by_date.items(), key=lambda x: x[0]):
            n = (
                self.db.query(SvDailyOperation)
                .filter(SvDailyOperation.report_id == rep.report_id)
                .count()
            )
            summary["daily_summaries"].append(
                {
                    "report_date": rd.isoformat(),
                    "operations_count": n,
                }
            )

        return summary

    def _get_or_create_well(self, well_number: str, well_info: Dict, project_code: str) -> Well:
        well = self.db.query(Well).filter(Well.well_number == well_number).first()
        if not well:
            well = Well(
                well_number=well_number,
                project_code=project_code,
                company="ПАО Татнефть",
            )
            self.db.add(well)
            self.db.flush()
        return well

    def _get_or_create_wellbore(self, well: Well) -> Wellbore:
        wellbore = self.db.query(Wellbore).filter(
            Wellbore.well_id == well.well_id,
            Wellbore.wellbore_number == "main",
        ).first()
        if not wellbore:
            wellbore = Wellbore(well_id=well.well_id, wellbore_number="main")
            self.db.add(wellbore)
            self.db.flush()
        return wellbore

    def _register_file(self, file_path: str, well: Well) -> Optional[File]:
        filename = os.path.basename(file_path)
        existing = self.db.query(File).filter(
            File.file_name == filename,
            File.well_id == well.well_id,
        ).first()
        if existing:
            return existing
        try:
            file_size = os.path.getsize(file_path)
        except OSError:
            file_size = None
        f = File(
            file_name=filename,
            file_path=file_path,
            file_type="xlsx",
            category="supervision_journal_final",
            well_id=well.well_id,
            file_size_bytes=file_size,
            processing_status="processing",
        )
        self.db.add(f)
        self.db.flush()
        return f
