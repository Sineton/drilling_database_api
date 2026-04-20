"""
Supervisor journal xlsx parser service.

Parses structured Excel files (supervisor daily drilling reports) and
populates the sv_* tables in the database.

Real file layout (3189д.xlsx, 2442 rows × 54 cols):
  R9-11:   Title with well number
  R13-19:  Dates and time summary
  R21-29:  General info — contractors (col2=role, col22=company)
  R33-67:  Well passport table (col2=label, col35=value)
  R49-50:  Casing construction (multi-line cells, col35=diam, col45=depth)
  R51-60:  Equipment
  R62-67:  Wellbore data
  R68-71:  Plan/fact timing
  R73:     Construction start
  R75+:    Daily blocks — repeating structure per day
  R~2294:  Construction completion
  R~2296-2301: NPV summary totals
  R~2336-2345: NPV detailed balance
  R~2410-2442: Final report + evaluations
"""
import re
import logging
from datetime import datetime, date, time, timedelta
from typing import Optional, Dict, Any, List, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from ..models import (
    Well, Wellbore, File,
    SvDailyReport, SvDailyOperation, SvBhaRun,
    SvDrillingRegime, SvMudAccounting, SvChemicalReagent,
    SvNpvBalance, SvContractor, SvWellConstruction,
    SvRigEquipment, SvConstructionTiming,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Anomaly detection
# ---------------------------------------------------------------------------
ANOMALY_PATTERNS: List[Dict[str, Any]] = [
    {"pattern": r"НПВ", "category": "НПВ", "severity": 2, "flag": "npv"},
    {"pattern": r"прихват", "category": "НПВ", "severity": 3, "flag": "stuck_pipe"},
    {"pattern": r"затяж[кеи]", "category": "осложнение", "severity": 2, "flag": "drag"},
    {"pattern": r"посадк[аие]", "category": "осложнение", "severity": 2, "flag": "set_down"},
    {"pattern": r"забити[ей]?\s*ВЗД", "category": "НПВ", "severity": 2, "flag": "mud_motor_plugged"},
    {"pattern": r"рост\s*(да)?[лв]ения", "category": "осложнение", "severity": 2, "flag": "pressure_rise"},
    {"pattern": r"поглощени[ея]", "category": "осложнение", "severity": 2, "flag": "absorption"},
    {"pattern": r"потер[яи]\s*циркуляции", "category": "осложнение", "severity": 3, "flag": "circulation_loss"},
    {"pattern": r"геолог\w*\s*осложнени", "category": "осложнение", "severity": 3, "flag": "geological_complication"},
    {"pattern": r"отказ", "category": "НПВ", "severity": 2, "flag": "equipment_failure"},
    {"pattern": r"[Дд]оп\.?\s*работ", "category": "доп_работы", "severity": 1, "flag": "additional_work"},
    {"pattern": r"[Рр]емонт", "category": "ремонт", "severity": 2, "flag": "repair"},
    {"pattern": r"засорени", "category": "НПВ", "severity": 2, "flag": "plugging"},
]

OPERATION_CATEGORY_PATTERNS: List[Tuple[str, str]] = [
    (r"[Бб]урени[ея]", "бурение"),
    (r"[Сс]пуск\s", "СПО"),
    (r"[Пп]одъ[её]м", "СПО"),
    (r"СПО", "СПО"),
    (r"[Нн]аращивани", "наращивание"),
    (r"[Пп]ромывк", "промывка"),
    (r"[Пп]роработк", "проработка"),
    (r"[Цц]ементирован", "цементирование"),
    (r"ГИС|[Кк]аротаж", "ГИС"),
    (r"НПВ", "НПВ"),
    (r"[Дд]оп\.?\s*работ", "доп_работы"),
    (r"[Рр]емонт", "ремонт"),
    (r"ПЗР|подготови", "ПЗР"),
    (r"ВМР", "ВМР"),
    (r"[Мм]онтаж.*ПВО|ПВО", "монтаж_ПВО"),
    (r"[Оо]жидани", "ожидание"),
]

DRILLING_PARAM_PATTERNS = {
    "G_ton": r"G\s*=\s*([\d.,\-]+)\s*т",
    "Q_l_s": r"Q\s*=\s*([\d.,\-]+)\s*л/с",
    "P_atm": r"[PРр]\s*=\s*([\d.,\-]+)\s*(?:атм|кгс)",
    "N_rpm": r"[Nn]\s*=\s*([\d.,\-]+)\s*об",
    "M_kNm": r"[Мм]всп?\s*=\s*([\d.,\-]+)",
}


# ---------------------------------------------------------------------------
# Cell value helpers
# ---------------------------------------------------------------------------

def _cell(ws: Worksheet, row: int, col: int) -> Any:
    v = ws.cell(row=row, column=col).value
    return v


def _float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        if isinstance(val, str):
            val = val.replace(",", ".").replace("\xa0", "").replace(" ", "").strip()
            if not val or val == "-" or val == "/":
                return None
        return float(val)
    except (ValueError, TypeError):
        return None


def _int(val) -> Optional[int]:
    f = _float(val)
    return int(f) if f is not None else None


def _text(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("none", "nan", "-") else None


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_time(val) -> Optional[time]:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    s = str(val).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%H.%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    m = re.match(r"(\d{1,2})[:\.](\d{2})", s)
    if m:
        try:
            return time(int(m.group(1)), int(m.group(2)))
        except ValueError:
            pass
    return None


def _parse_duration_minutes(text: str) -> Optional[int]:
    if not text:
        return None
    total = 0
    h_match = re.search(r"(\d+)\s*час", text)
    m_match = re.search(r"(\d+)\s*мин", text)
    if h_match:
        total += int(h_match.group(1)) * 60
    if m_match:
        total += int(m_match.group(1))
    if total == 0:
        f = _float(text)
        if f is not None:
            total = int(f * 60) if f < 24 else int(f)
    return total if total > 0 else None


def _classify_operation(description: str) -> Optional[str]:
    if not description:
        return None
    for pattern, category in OPERATION_CATEGORY_PATTERNS:
        if re.search(pattern, description, re.IGNORECASE):
            return category
    return "прочее"


def _detect_anomalies(description: str) -> Tuple[Dict, int, bool, bool, bool]:
    flags = {}
    max_severity = 0
    is_npv = False
    is_complication = False
    is_repair = False
    if not description:
        return flags, 0, False, False, False
    for ap in ANOMALY_PATTERNS:
        if re.search(ap["pattern"], description, re.IGNORECASE):
            flags[ap["flag"]] = True
            max_severity = max(max_severity, ap["severity"])
            if ap["category"] == "НПВ":
                is_npv = True
            if ap["category"] == "осложнение":
                is_complication = True
            if ap["category"] == "ремонт":
                is_repair = True
    return flags, max_severity, is_npv, is_complication, is_repair


def _extract_params(description: str) -> Optional[Dict]:
    if not description:
        return None
    params = {}
    for key, pattern in DRILLING_PARAM_PATTERNS.items():
        m = re.search(pattern, description, re.IGNORECASE)
        if m:
            params[key] = m.group(1)
    depth_m = re.search(r"(?:гл|глубин[аеы]|на)\s*\.?\s*(\d+(?:[.,]\d+)?)\s*м", description, re.IGNORECASE)
    if depth_m:
        params["depth_m"] = depth_m.group(1)
    return params if params else None


def _extract_depth_range(description: str) -> Tuple[Optional[float], Optional[float]]:
    if not description:
        return None, None
    text = description.replace("\xa0", " ")

    range_patterns = [
        # "в инт. 24,23-403м", "в интервале 475-518м"
        r"(?:в\s*)?(?:инт(?:ервал[еа]?)?\.?|интервал[еа]?)\s*(\d+(?:[.,]\d+)?)\s*[-–—]\s*(\d+(?:[.,]\d+)?)\s*м",
        # "от 475 до 868", "от 475м до 868м"
        r"от\s+(\d+(?:[.,]\d+)?)\s*м?\s*до\s+(\d+(?:[.,]\d+)?)\s*м?",
        # Generic depth interval with explicit meters suffix: "385-300м"
        r"\b(\d+(?:[.,]\d+)?)\s*[-–—]\s*(\d+(?:[.,]\d+)?)\s*м\b",
    ]
    for pattern in range_patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            return _float(m.group(1)), _float(m.group(2))

    # Single depth reference: "на гл. 403м" -> use same value for from/to
    m = re.search(r"(?:на\s+гл\.?|на\s+глубине|глубин[аеы]?|до\s+гл\.?)\s*(\d+(?:[.,]\d+)?)\s*м", text, re.IGNORECASE)
    if m:
        d = _float(m.group(1))
        return d, d

    return None, None


def _extract_numeric(val) -> Optional[float]:
    """Extract first number from text like '20 сут.' or '1,21 сут.'"""
    if val is None:
        return None
    f = _float(val)
    if f is not None:
        return f
    s = str(val).strip()
    m = re.search(r"([\d]+[.,]?\d*)", s.replace("\xa0", "").replace(" ", ""))
    if m:
        return _float(m.group(1))
    return None


def _parse_range(val) -> Tuple[Optional[float], Optional[float]]:
    if val is None:
        return None, None
    s = str(val).strip()
    m = re.match(r"([\d.,]+)\s*[-–—]\s*([\d.,]+)", s)
    if m:
        return _float(m.group(1)), _float(m.group(2))
    f = _float(s)
    return f, f


class SvJournalParserService:
    """Parses supervisor journal xlsx and imports data into the database."""

    def __init__(self, db: Session):
        self.db = db
        self.warnings: List[str] = []
        self.errors: List[str] = []

    # ------------------------------------------------------------------
    # Public: parse preview
    # ------------------------------------------------------------------
    def parse_preview(self, file_path: str) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        well_info = self._parse_well_passport(ws)
        daily_blocks = self._find_daily_blocks(ws)

        date_range = None
        if daily_blocks:
            dates = [b["date"] for b in daily_blocks if b.get("date")]
            if dates:
                date_range = {"from": min(dates).isoformat(), "to": max(dates).isoformat()}

        npv_rows = self._find_npv_rows(ws)

        return {
            "success": True,
            "file_info": {
                "filename": file_path.split("\\")[-1].split("/")[-1],
                "sheet": ws.title,
                "total_rows": ws.max_row,
                "total_cols": ws.max_column,
            },
            "well_info": well_info,
            "daily_blocks_count": len(daily_blocks),
            "date_range": date_range,
            "npv_count": len(npv_rows),
            "warnings": self.warnings,
        }

    # ------------------------------------------------------------------
    # Public: full import
    # ------------------------------------------------------------------
    def import_journal(
        self,
        file_path: str,
        project_code: str = "pao-tatneft",
        well_number_override: Optional[str] = None,
        dry_run: bool = False,
    ) -> Dict[str, Any]:
        self.warnings = []
        self.errors = []

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        well_info = self._parse_well_passport(ws)
        well_num = well_number_override or well_info.get("well_number")
        if not well_num:
            raise ValueError("Не удалось определить номер скважины. Укажите well_number.")

        summary = {
            "well_number": well_num,
            "daily_reports_created": 0,
            "operations_created": 0,
            "bha_runs_created": 0,
            "drilling_regimes_created": 0,
            "mud_accounting_created": 0,
            "chemical_reagents_created": 0,
            "npv_records_created": 0,
            "contractors_created": 0,
            "construction_items_created": 0,
            "equipment_created": 0,
            "timing_records_created": 0,
            "mud_properties_created": 0,
        }

        if dry_run:
            daily_blocks = self._find_daily_blocks(ws)
            summary["daily_reports_created"] = len(daily_blocks)
            summary["warnings"] = self.warnings
            summary["errors"] = self.errors
            summary["daily_summaries"] = []
            return summary

        well = self._get_or_create_well(well_num, well_info, project_code)
        summary["well_id"] = well.well_id
        wellbore = self._get_or_create_wellbore(well, well_info)
        summary["wellbore_id"] = wellbore.wellbore_id
        file_record = self._register_file(file_path, well)

        self._update_well_passport(well, well_info)
        self._update_wellbore_passport(wellbore, well_info)

        # Construction
        for item in self._parse_well_construction(ws):
            self.db.add(SvWellConstruction(wellbore_id=wellbore.wellbore_id, **item))
            summary["construction_items_created"] += 1

        # Equipment
        equip = self._parse_rig_equipment(ws)
        if equip:
            self.db.add(SvRigEquipment(wellbore_id=wellbore.wellbore_id, **equip))
            summary["equipment_created"] += 1

        # Timing
        for item in self._parse_construction_timing(ws):
            self.db.add(SvConstructionTiming(wellbore_id=wellbore.wellbore_id, **item))
            summary["timing_records_created"] += 1

        # Contractors
        for c in self._parse_contractors(ws):
            existing = self.db.query(SvContractor).filter(
                SvContractor.wellbore_id == wellbore.wellbore_id,
                SvContractor.role == c["role"]
            ).first()
            if not existing:
                self.db.add(SvContractor(wellbore_id=wellbore.wellbore_id, **c))
                summary["contractors_created"] += 1

        self.db.flush()

        # Daily blocks
        daily_blocks = self._find_daily_blocks(ws)
        daily_summaries = []

        for block in daily_blocks:
            report_date = block["date"]
            if not report_date:
                self.warnings.append(f"Пропущен блок строка {block['row_start']}: нет даты")
                continue

            existing = self.db.query(SvDailyReport).filter(
                SvDailyReport.wellbore_id == wellbore.wellbore_id,
                SvDailyReport.report_date == report_date,
            ).first()
            if existing:
                self.warnings.append(f"Отчёт за {report_date} уже существует, пропуск")
                continue

            rd = self._parse_daily_block(ws, block)

            report = SvDailyReport(
                wellbore_id=wellbore.wellbore_id,
                report_date=report_date,
                construction_stage=rd.get("construction_stage"),
                interval_from_m=rd.get("interval_from_m"),
                interval_to_m=rd.get("interval_to_m"),
                current_depth_m=rd.get("current_depth_m"),
                penetration_m=rd.get("penetration_m"),
                drilling_time_h=rd.get("drilling_time_h"),
                rop_plan=rd.get("rop_plan"),
                rop_fact=rd.get("rop_fact"),
                cum_drilling_time_plan_h=rd.get("cum_drilling_time_plan_h"),
                cum_drilling_time_fact_h=rd.get("cum_drilling_time_fact_h"),
                cum_penetration_m=rd.get("cum_penetration_m"),
                cum_avg_rop=rd.get("cum_avg_rop"),
                drilling_comment=rd.get("drilling_comment"),
                source_file_id=file_record.file_id if file_record else None,
                source_row_start=block["row_start"],
            )
            self.db.add(report)
            self.db.flush()
            summary["daily_reports_created"] += 1

            ops_count = 0
            for op_data in rd.get("operations", []):
                self.db.add(SvDailyOperation(report_id=report.report_id, **op_data))
                ops_count += 1
            summary["operations_created"] += ops_count

            for bha_data in rd.get("bha_runs", []):
                self.db.add(SvBhaRun(report_id=report.report_id, **bha_data))
                summary["bha_runs_created"] += 1

            for reg_data in rd.get("drilling_regimes", []):
                self.db.add(SvDrillingRegime(report_id=report.report_id, **reg_data))
                summary["drilling_regimes_created"] += 1

            for mud_data in rd.get("mud_accounting", []):
                self.db.add(SvMudAccounting(report_id=report.report_id, **mud_data))
                summary["mud_accounting_created"] += 1

            for chem_data in rd.get("chemical_reagents", []):
                self.db.add(SvChemicalReagent(report_id=report.report_id, **chem_data))
                summary["chemical_reagents_created"] += 1

            daily_summaries.append({
                "report_date": report_date.isoformat(),
                "construction_stage": rd.get("construction_stage"),
                "current_depth_m": rd.get("current_depth_m"),
                "penetration_m": rd.get("penetration_m"),
                "operations_count": ops_count,
                "bha_count": len(rd.get("bha_runs", [])),
                "npv_found": any(op.get("is_npv") for op in rd.get("operations", [])),
            })

        # NPV balance
        for npv_data in self._parse_npv_balance(ws):
            self.db.add(SvNpvBalance(
                wellbore_id=wellbore.wellbore_id,
                source_file_id=file_record.file_id if file_record else None,
                **npv_data,
            ))
            summary["npv_records_created"] += 1

        self.db.commit()

        if file_record:
            file_record.processing_status = "completed"
            self.db.commit()

        summary["warnings"] = self.warnings
        summary["errors"] = self.errors
        summary["daily_summaries"] = daily_summaries
        return summary

    # ==================================================================
    # WELL PASSPORT (rows 1–70)
    # ==================================================================

    def _parse_well_passport(self, ws: Worksheet) -> Dict[str, Any]:
        info: Dict[str, Any] = {}

        # Title rows (9-11): extract well number
        for row in range(8, 15):
            val = _text(_cell(ws, row, 2))
            if val:
                m = re.search(r"СКВАЖИН[ЫА]\s*[№#]?\s*(\S+)", val, re.IGNORECASE)
                if m:
                    info["well_number"] = m.group(1).rstrip(".")

        # Dates and time summary (rows 13-19)
        for row in range(13, 20):
            label = _text(_cell(ws, row, 2))
            val19 = _text(_cell(ws, row, 19))
            if not label:
                continue
            ll = label.lower()
            if "начало бурения" in ll:
                info["drilling_start_date"] = _parse_date(val19)
            elif "конец бурения" in ll or "окончание" in ll:
                info["drilling_end_date"] = _parse_date(val19)
            elif "календарное время" in ll:
                info["calendar_days"] = _extract_numeric(val19)
            elif "внеплановые" in ll:
                info["unplanned_days"] = _extract_numeric(val19)
            elif "нпв" in ll and "ремонт" in ll:
                info["npv_repair_days"] = _extract_numeric(val19)
            elif "простой" in ll:
                info["idle_days"] = _extract_numeric(val19)
            elif "осложнение" in ll:
                info["complication_days"] = _extract_numeric(val19)

        # Contractors (rows 21-30): col 2=role, col 22=company
        for row in range(20, 35):
            label = _text(_cell(ws, row, 2))
            value = _text(_cell(ws, row, 22))
            if label and value:
                info.setdefault("contractors", []).append({
                    "label": label, "company": value
                })
                ll = label.lower()
                if "заказчик" in ll:
                    info["customer"] = value

        # Passport table (rows 33-67): col 2=label, col 35=value
        for row in range(33, 68):
            label = _text(_cell(ws, row, 2))
            value35 = _cell(ws, row, 35)
            value45 = _cell(ws, row, 45)
            if not label:
                continue
            ll = label.lower()
            v = _text(value35)

            if "площадь" in ll or "месторождение" in ll:
                info["field_name"] = v
            elif "номер скважины" in ll:
                if v:
                    info["well_number"] = v
            elif "групповой проект" in ll:
                info["group_project"] = v
            elif "назначение скважины" in ll:
                info["well_purpose"] = v
            elif "проектный горизонт" in ll:
                info["target_horizon"] = v
            elif "продуктивный пласт" in ll:
                info["productive_layer"] = v
            elif "альтитуда" in ll:
                info["altitude_rotor_m"] = _float(value35)
            elif "абсолютная отметка" in ll and "кровл" in ll:
                info["abs_mark_top_m"] = _float(value35)
            elif "абсолютная отметка" in ll and "забо" in ll:
                info["abs_mark_bottom_m"] = _float(value35)
            elif "по верикали" in ll or "по вертикали" in ll:
                parent = _text(_cell(ws, row - 1, 2)) or ""
                if "кровл" in parent.lower():
                    info["design_depth_vertical_m"] = _float(value35)
                elif "забо" in parent.lower():
                    info["wellbore_design_depth_vertical_m"] = _float(value35)
                else:
                    info["design_depth_vertical_m"] = _float(value35)
            elif "по стволу" in ll:
                parent = _text(_cell(ws, row - 2, 2)) or _text(_cell(ws, row - 1, 2)) or ""
                if "забо" in parent.lower():
                    info["wellbore_design_depth_md_m"] = _float(value35)
                else:
                    info["design_depth_md_m"] = _float(value35)
            elif "магнитный азимут" in ll:
                info["magnetic_azimuth"] = _float(value35)
            elif "проектное смещение" in ll:
                info["design_offset_m"] = _float(value35)
            elif "радиус" in ll and "допуск" in ll:
                info["tolerance_radius_m"] = _float(value35)
            elif "смещение на забой" in ll:
                info["offset_to_bottom_m"] = _float(value35)
            elif "азимут" in ll and "забой" in ll:
                info["azimuth_to_bottom"] = _float(value35)
            elif "тип буровой" in ll:
                info["rig_type"] = v
            elif "вид монтажа" in ll:
                info["mounting_type"] = v
            elif "оснастка" in ll or "талевой" in ll:
                info["talev_system"] = v
            elif "буровой насос" in ll or ("насос" in ll and "тип" in ll):
                info["pump_type"] = v
            elif "вибросит" in ll:
                info["shaker_type"] = v
            elif "гидроциклон" in ll:
                info["hydrocyclone_type"] = v
            elif "амбар" in ll:
                info["pit_description"] = v
            elif "ёмкост" in ll or "емкост" in ll:
                info["tank_system"] = v

        # Supervisor name (end of file)
        for row in range(max(1, ws.max_row - 10), ws.max_row + 1):
            for col in range(1, 15):
                val = _text(_cell(ws, row, col))
                if val and "супервайзер" in val.lower():
                    sv_name = _text(_cell(ws, row, 41)) or _text(_cell(ws, row, col + 10))
                    if sv_name:
                        info.setdefault("supervisors", []).append(sv_name)

        return info

    # ==================================================================
    # FIND DAILY BLOCKS
    # ==================================================================

    def _find_daily_blocks(self, ws: Worksheet) -> List[Dict]:
        """
        Each day starts with "Этап строительства скважины:" line,
        then a date row (e.g., "01.03.2024" in col 2),
        then "ПОКАЗАТЕЛИ БУРЕНИЯ" header.
        """
        blocks = []
        max_row = ws.max_row
        last_stage = None

        row = 60
        while row <= max_row:
            val_c2 = _text(_cell(ws, row, 2))

            # Detect stage header
            if val_c2 and "этап строительства" in val_c2.lower():
                stage_text = _text(_cell(ws, row, 15)) or _text(_cell(ws, row, 10))
                if stage_text:
                    last_stage = stage_text
                row += 1
                continue

            # Detect date
            d = _parse_date(val_c2)
            if d:
                if not blocks or d != blocks[-1].get("date"):
                    blocks.append({
                        "row_start": row,
                        "date": d,
                        "construction_stage": last_stage,
                    })
            row += 1

        # Set row_end
        for i in range(len(blocks) - 1):
            blocks[i]["row_end"] = blocks[i + 1]["row_start"] - 1
        if blocks:
            end_marker = self._find_completion_row(ws)
            blocks[-1]["row_end"] = end_marker if end_marker else min(blocks[-1]["row_start"] + 200, max_row)

        return blocks

    def _find_completion_row(self, ws: Worksheet) -> Optional[int]:
        for row in range(ws.max_row - 200, ws.max_row + 1):
            if row < 1:
                continue
            val = _text(_cell(ws, row, 2))
            if val and "строительство скважины завершено" in val.lower():
                return row
        return None

    # ==================================================================
    # PARSE SINGLE DAILY BLOCK
    # ==================================================================

    def _parse_daily_block(self, ws: Worksheet, block: Dict) -> Dict[str, Any]:
        rs = block["row_start"]
        re_row = block["row_end"]
        result: Dict[str, Any] = {}

        result["construction_stage"] = block.get("construction_stage")
        if result["construction_stage"]:
            d_from, d_to = _extract_depth_range(result["construction_stage"])
            result["interval_from_m"] = d_from
            result["interval_to_m"] = d_to

        # Scan for section headers
        sec_indicators = None
        sec_operations = None
        sec_mud_props = None
        sec_mud_acct = None
        sec_chem = None
        sec_bha = None
        sec_regime = None

        for row in range(rs, min(re_row + 1, ws.max_row + 1)):
            val = _text(_cell(ws, row, 5)) or _text(_cell(ws, row, 2))
            if not val:
                continue
            vl = val.lower()

            if "показатели бурения" in vl:
                sec_indicators = row
            elif "выполненные" in vl and "операци" in vl:
                sec_operations = row
            elif "буровые растворы" in vl or "буровой раствор" in vl.replace(" ", ""):
                if "учет" not in vl and "учёт" not in vl:
                    sec_mud_props = row
            elif "учет бурового" in vl or "учёт бурового" in vl:
                sec_mud_acct = row
            elif "хим" in vl and "реагент" in vl:
                sec_chem = row
            elif "конструкция низа" in vl or "кнбк" in vl:
                sec_bha = row
            elif "режим" in vl and "бурен" in vl:
                sec_regime = row

        # Drilling indicators
        if sec_indicators:
            self._parse_drilling_indicators(ws, sec_indicators, result)

        # Operations
        if sec_operations:
            end = self._section_end(sec_operations, [sec_mud_props, sec_bha, sec_regime, sec_mud_acct, sec_chem, re_row])
            result["operations"] = self._parse_operations(ws, sec_operations, end)
        else:
            result["operations"] = []

        # BHA
        if sec_bha:
            end = self._section_end(sec_bha, [sec_regime, sec_mud_props, sec_mud_acct, sec_chem, re_row])
            result["bha_runs"] = self._parse_bha(ws, sec_bha, end)
        else:
            result["bha_runs"] = []

        # Drilling regime
        if sec_regime:
            end = self._section_end(sec_regime, [sec_mud_props, sec_mud_acct, sec_chem, re_row])
            result["drilling_regimes"] = self._parse_drilling_regime(ws, sec_regime, end)
        else:
            result["drilling_regimes"] = []

        # Mud accounting
        if sec_mud_acct:
            end = self._section_end(sec_mud_acct, [sec_chem, re_row])
            result["mud_accounting"] = self._parse_mud_accounting(ws, sec_mud_acct, end)
        else:
            result["mud_accounting"] = []

        # Chemical reagents
        if sec_chem:
            result["chemical_reagents"] = self._parse_chemical_reagents(ws, sec_chem, re_row)
        else:
            result["chemical_reagents"] = []

        return result

    def _section_end(self, start: int, candidates: List[Optional[int]]) -> int:
        valid = [c for c in candidates if c is not None and c > start]
        return min(valid) if valid else start + 100

    # ------------------------------------------------------------------
    # DRILLING INDICATORS (row after "ПОКАЗАТЕЛИ БУРЕНИЯ")
    # ------------------------------------------------------------------

    def _parse_drilling_indicators(self, ws: Worksheet, header_row: int, result: Dict):
        """
        Data row: header_row + 2 (after column labels + plan/fact row)
        col 5=depth, col 11=penetration, col 18=drill_time,
        col 24=rop_plan, col 28=rop_fact, col 32=cum_plan, col 36=cum_fact,
        col 40=cum_penetration, col 45=cum_avg_rop, col 50=comment
        """
        data_row = header_row + 3
        result["current_depth_m"] = _float(_cell(ws, data_row, 5))
        result["penetration_m"] = _float(_cell(ws, data_row, 11))
        result["drilling_time_h"] = _float(_cell(ws, data_row, 18))
        result["rop_plan"] = _float(_cell(ws, data_row, 24))
        result["rop_fact"] = _float(_cell(ws, data_row, 28))
        result["cum_drilling_time_plan_h"] = _float(_cell(ws, data_row, 32))
        result["cum_drilling_time_fact_h"] = _float(_cell(ws, data_row, 36))
        result["cum_penetration_m"] = _float(_cell(ws, data_row, 40))
        result["cum_avg_rop"] = _float(_cell(ws, data_row, 45))
        result["drilling_comment"] = _text(_cell(ws, data_row, 50))

    # ------------------------------------------------------------------
    # OPERATIONS
    # ------------------------------------------------------------------

    def _parse_operations(self, ws: Worksheet, header_row: int, section_end: int) -> List[Dict]:
        """
        Starts after 2 header rows (col names + From/To/Time).
        col 2=seq, col 5=time_from, col 9=time_to, col 13=duration,
        col 19=description, col 38=operation type, col 49=comment
        """
        operations = []
        data_start = header_row + 3

        for row in range(data_start, min(section_end, ws.max_row + 1)):
            seq_val = _int(_cell(ws, row, 2))
            desc = _text(_cell(ws, row, 19))
            op_type = _text(_cell(ws, row, 38))

            if not desc and not op_type:
                if operations:
                    break
                continue

            if not desc:
                desc = op_type or ""

            tf = _parse_time(_cell(ws, row, 5))
            tt = _parse_time(_cell(ws, row, 9))
            dur_text = _text(_cell(ws, row, 13))
            dur_min = _parse_duration_minutes(dur_text) if dur_text else None
            comment = _text(_cell(ws, row, 49))

            if not dur_min and tf and tt:
                t1 = datetime.combine(date.today(), tf)
                t2 = datetime.combine(date.today(), tt)
                if t2 < t1:
                    t2 += timedelta(days=1)
                dur_min = int((t2 - t1).total_seconds() / 60)

            full_desc = desc
            if comment:
                full_desc = f"{desc} ({comment})"

            category = _classify_operation(full_desc)
            flags, severity, is_npv, is_complication, is_repair = _detect_anomalies(full_desc)
            params = _extract_params(full_desc)
            depth_from, depth_to = _extract_depth_range(full_desc)

            operations.append({
                "sequence_number": seq_val or len(operations) + 1,
                "time_from": tf,
                "time_to": tt,
                "duration_text": dur_text,
                "duration_minutes": dur_min,
                "description": full_desc,
                "operation_category": category,
                "is_npv": is_npv,
                "is_complication": is_complication,
                "is_repair": is_repair,
                "extracted_params": params,
                "depth_from_m": depth_from,
                "depth_to_m": depth_to,
                "anomaly_flags": flags if flags else None,
                "anomaly_severity": severity,
                "source_row": row,
            })

        return operations

    # ------------------------------------------------------------------
    # BHA
    # ------------------------------------------------------------------

    def _parse_bha(self, ws: Worksheet, header_row: int, section_end: int) -> List[Dict]:
        bha_list = []
        current_status = "fact"

        for row in range(header_row, min(section_end, ws.max_row + 1)):
            for col in range(1, 10):
                val = _text(_cell(ws, row, col))
                if val:
                    vl = val.lower()
                    if "план" in vl and "факт" not in vl:
                        current_status = "plan"
                    elif "факт" in vl:
                        current_status = "fact"

            num_val = _int(_cell(ws, row, 2))
            desc = _text(_cell(ws, row, 5))

            if not desc:
                parts = []
                for c in range(3, 50):
                    t = _text(_cell(ws, row, c))
                    if t and len(t) > 5:
                        parts.append(t)
                if parts:
                    desc = "\n".join(parts)

            if desc and len(desc) > 10:
                bit_type, bit_size, motor_type = self._parse_bha_components(desc)
                bha_list.append({
                    "bha_number": num_val,
                    "status": current_status,
                    "description": desc,
                    "bit_type": bit_type,
                    "bit_size_mm": bit_size,
                    "motor_type": motor_type,
                    "source_row": row,
                })

        return bha_list

    def _parse_bha_components(self, desc: str) -> Tuple[Optional[str], Optional[float], Optional[str]]:
        bit_type = None
        bit_size = None
        motor_type = None

        bit_match = re.search(r"(\d{3}[.,]\d+)\s*PDC|PDC\s*[(\[]?([\d.,]+)", desc, re.IGNORECASE)
        if bit_match:
            size_str = bit_match.group(1) or bit_match.group(2)
            bit_size = _float(size_str)
            first_line = desc.split("\n")[0] if "\n" in desc else desc[:80]
            bit_type = first_line

        motor_match = re.search(r"(ВЗД[-\s]*\d+|ДР[-\s]*\d+[\s/\d]*)", desc, re.IGNORECASE)
        if motor_match:
            motor_type = motor_match.group(1).strip()

        return bit_type, bit_size, motor_type

    # ------------------------------------------------------------------
    # DRILLING REGIME
    # ------------------------------------------------------------------

    def _parse_drilling_regime(self, ws: Worksheet, header_row: int, section_end: int) -> List[Dict]:
        regimes = []
        for row in range(header_row + 1, min(section_end, ws.max_row + 1)):
            status = "fact"
            label = _text(_cell(ws, row, 2))
            if label:
                if "план" in label.lower():
                    status = "plan"

            wob_val = _cell(ws, row, 5)
            if wob_val is None:
                continue

            wob_min, wob_max = _parse_range(wob_val)
            rpm_min, rpm_max = _parse_range(_cell(ws, row, 11))
            p_min, p_max = _parse_range(_cell(ws, row, 18))
            flow = _float(_cell(ws, row, 24))
            dp = _float(_cell(ws, row, 29))
            pumps = _int(_cell(ws, row, 33))
            liner = _float(_cell(ws, row, 37))
            notes = _text(_cell(ws, row, 42))

            if wob_min is not None or rpm_min is not None or p_min is not None:
                regimes.append({
                    "status": status,
                    "wob_min_ton": wob_min, "wob_max_ton": wob_max,
                    "rpm_min": rpm_min, "rpm_max": rpm_max,
                    "pressure_min": p_min, "pressure_max": p_max,
                    "flow_rate_l_s": flow, "delta_p": dp,
                    "pump_count": pumps, "liner_diameter_mm": liner,
                    "notes": notes, "source_row": row,
                })
        return regimes

    # ------------------------------------------------------------------
    # MUD ACCOUNTING (col 2=label, col 30=value)
    # ------------------------------------------------------------------

    def _parse_mud_accounting(self, ws: Worksheet, header_row: int, section_end: int) -> List[Dict]:
        accounts = []
        current_type = None
        current_acc = None

        for row in range(header_row, min(section_end, ws.max_row + 1)):
            label = _text(_cell(ws, row, 2))
            val30 = _cell(ws, row, 30)
            val24 = _text(_cell(ws, row, 24))

            if not label:
                continue
            ll = label.lower()

            # Detect mud type header: "Буровой раствор" row with type in col 30
            if "буровой раствор" in ll and val30:
                type_name = _text(val30)
                if type_name and "показатель" not in type_name.lower():
                    if current_acc:
                        accounts.append(current_acc)
                    current_type = type_name
                    current_acc = {"mud_type": current_type, "source_row": row}
                continue

            if not current_acc:
                continue

            v = _float(val30)

            if "на начало" in ll:
                current_acc["volume_start"] = v
            elif "приготовлено" in ll:
                current_acc["volume_prepared"] = v
            elif "утяжеление" in ll:
                current_acc["volume_weighted"] = v
            elif "завезено" in ll:
                current_acc["volume_delivered"] = v
            elif "вывезено" in ll:
                current_acc["volume_exported"] = v
            elif "утилизир" in ll:
                current_acc["volume_disposed"] = v
            elif "увеличено" in ll:
                current_acc["volume_increased"] = v
            elif "общие потери" in ll:
                current_acc["total_losses"] = v
            elif "на поверхности" in ll and "потери" not in ll:
                current_acc["surface_losses"] = v
            elif "системе очистки" in ll:
                current_acc["cleaning_losses"] = v
            elif "при спо" in ll.replace(" ", ""):
                current_acc["spo_losses"] = v
            elif "розлив" in ll:
                current_acc["spill_losses"] = v
            elif "чистке емкостей" in ll:
                current_acc["tank_cleaning"] = v
            elif "сброс в амбар" in ll:
                current_acc["pit_discharge"] = v
            elif "зона перемешивания" in ll:
                current_acc["cement_zone"] = v
            elif "переходе на раствор" in ll:
                current_acc["mud_transition"] = v
            elif "потери в скважине" in ll:
                current_acc["downhole_losses"] = v
            elif "поглощение" in ll:
                current_acc["absorption"] = v
            elif "намыве" in ll:
                current_acc["washout"] = v
            elif "остаток в стволе" in ll:
                current_acc["wellbore_remain"] = v
            elif "естественная фильтрация" in ll:
                current_acc["filtration"] = v
            elif "прокачка" in ll:
                current_acc["circulation_pump"] = v
            elif "остаток" in ll and "скважине" in ll:
                current_acc["volume_remaining"] = v

        if current_acc:
            accounts.append(current_acc)

        return accounts

    # ------------------------------------------------------------------
    # CHEMICAL REAGENTS (col 2=name, col 18=unit, col 22+=values)
    # ------------------------------------------------------------------

    def _parse_chemical_reagents(self, ws: Worksheet, header_row: int, section_end: int) -> List[Dict]:
        reagents = []
        # Data starts after 3 header rows (name row, sub-headers, sub-sub-headers)
        data_start = header_row + 4

        for row in range(data_start, min(section_end, ws.max_row + 1)):
            name = _text(_cell(ws, row, 2))
            if not name or len(name) < 3:
                continue
            if name == ".":
                continue
            nl = name.lower()
            if any(kw in nl for kw in ["наименование", "итого", "всего"]):
                continue

            unit = _text(_cell(ws, row, 18))
            received = _float(_cell(ws, row, 22))
            used_prep = _float(_cell(ws, row, 26))
            used_treat = _float(_cell(ws, row, 31))
            used_regen = _float(_cell(ws, row, 35))
            exported = _float(_cell(ws, row, 40))
            remaining = _float(_cell(ws, row, 44))

            if received is not None or remaining is not None or used_prep is not None:
                reagents.append({
                    "reagent_name": name,
                    "unit": unit,
                    "total_received": received,
                    "used_preparation": used_prep,
                    "used_treatment": used_treat,
                    "used_regeneration": used_regen,
                    "exported": exported,
                    "remaining": remaining,
                    "source_row": row,
                })

        return reagents

    # ==================================================================
    # NPV BALANCE (rows ~2336+)
    # ==================================================================

    def _find_npv_rows(self, ws: Worksheet) -> List[int]:
        npv_rows = []
        for row in range(max(1, ws.max_row - 200), ws.max_row + 1):
            val = _text(_cell(ws, row, 2))
            if val and "баланс" in val.lower() and ("нпв" in val.lower() or "непроизводительн" in val.lower()):
                for data_row in range(row + 2, min(row + 50, ws.max_row + 1)):
                    date_val = _cell(ws, data_row, 2)
                    desc_val = _text(_cell(ws, data_row, 11))
                    if _parse_date(date_val) and desc_val:
                        npv_rows.append(data_row)
                    elif _text(date_val) and "общее время" in _text(date_val).lower():
                        break
                return npv_rows
        return npv_rows

    def _parse_npv_balance(self, ws: Worksheet) -> List[Dict]:
        items = []
        for row in self._find_npv_rows(ws):
            incident_date = _parse_date(_cell(ws, row, 2))
            description = _text(_cell(ws, row, 11))
            duration_hours = _float(_cell(ws, row, 28))
            responsible = _text(_cell(ws, row, 33))
            category = _text(_cell(ws, row, 42)) or "НеплановыеРаботы"
            operation_type = _text(_cell(ws, row, 48))

            if incident_date and description:
                items.append({
                    "incident_date": incident_date,
                    "description": description,
                    "duration_hours": duration_hours,
                    "responsible_party": responsible,
                    "category": category,
                    "operation_type": operation_type,
                    "source_row": row,
                })
        return items

    # ==================================================================
    # WELL CONSTRUCTION (row 49-50)
    # ==================================================================

    def _parse_well_construction(self, ws: Worksheet) -> List[Dict]:
        """
        Row 49: header ("Конструкция скважины", col35="Øнар, мм", col45="Hствола, м")
        Row 50: multi-line cell in col2 with types, col35 with diameters, col45 with depths
        """
        items = []
        for row in range(45, 55):
            val = _text(_cell(ws, row, 2))
            if not val:
                continue
            if "конструкция скважины" in val.lower():
                data_row = row + 1
                types_cell = _text(_cell(ws, data_row, 2))
                diams_cell = _text(_cell(ws, data_row, 35))
                depths_cell = _text(_cell(ws, data_row, 45))

                if types_cell:
                    types = [t.strip() for t in types_cell.split("\n") if t.strip()]
                    diams = [d.strip() for d in (diams_cell or "").split("\n")] if diams_cell else []
                    depths = [d.strip() for d in (depths_cell or "").split("\n")] if depths_cell else []

                    for i, ctype in enumerate(types):
                        items.append({
                            "casing_type": ctype,
                            "outer_diameter_mm": _float(diams[i]) if i < len(diams) else None,
                            "depth_m": _float(depths[i]) if i < len(depths) else None,
                        })
                break
        return items

    # ==================================================================
    # RIG EQUIPMENT (rows 51-60)
    # ==================================================================

    def _parse_rig_equipment(self, ws: Worksheet) -> Optional[Dict]:
        info = self._parse_well_passport.__func__  # already parsed
        # Re-read from passport info dict is cleaner; but let's scan directly
        equip: Dict[str, Any] = {}
        for row in range(48, 62):
            label = _text(_cell(ws, row, 2))
            val = _text(_cell(ws, row, 35))
            if not label:
                continue
            ll = label.lower()
            if "тип буровой" in ll:
                equip["rig_type"] = val
            elif "оснастка" in ll or "талевой" in ll:
                equip["talev_system"] = val
            elif "насос" in ll:
                equip["pump_type"] = val
            elif "вибросит" in ll:
                equip["shaker_type"] = val
            elif "гидроциклон" in ll:
                equip["hydrocyclone_type"] = val
            elif "амбар" in ll:
                equip["pit_description"] = val
            elif "ёмкост" in ll or "емкост" in ll:
                equip["tank_system"] = val
        return equip if equip else None

    # ==================================================================
    # CONSTRUCTION TIMING (rows 68-71)
    # ==================================================================

    def _parse_construction_timing(self, ws: Worksheet) -> List[Dict]:
        """
        Row 68: "Продолжительность строительства скважины"
        Row 69: col2=Интервал, col18=По проекту, col37=Фактически
        Row 70: col18=час, col28=сут, col37=час, col46=сут
        Row 71: Multi-line cell with interval names and values
        """
        items = []
        for row in range(65, 75):
            val = _text(_cell(ws, row, 2))
            if val and "продолжительность" in val.lower():
                data_row = row + 3
                names_cell = _text(_cell(ws, data_row, 2))
                plan_h_cell = _text(_cell(ws, data_row, 18))
                plan_d_cell = _text(_cell(ws, data_row, 28))
                fact_h_cell = _text(_cell(ws, data_row, 37))
                fact_d_cell = _text(_cell(ws, data_row, 46))

                if names_cell:
                    names = [n.strip() for n in names_cell.split("\n") if n.strip()]
                    plan_h_vals = [v.strip() for v in (plan_h_cell or "").split("\n")]
                    plan_d_vals = [v.strip() for v in (plan_d_cell or "").split("\n")]
                    fact_h_vals = [v.strip() for v in (fact_h_cell or "").split("\n")]
                    fact_d_vals = [v.strip() for v in (fact_d_cell or "").split("\n")]

                    for i, name in enumerate(names):
                        if not name or "нет данных" in name.lower():
                            continue
                        items.append({
                            "interval_name": name,
                            "plan_hours": _float(plan_h_vals[i]) if i < len(plan_h_vals) else None,
                            "plan_days": _float(plan_d_vals[i].lstrip("/")) if i < len(plan_d_vals) else None,
                            "fact_hours": _float(fact_h_vals[i]) if i < len(fact_h_vals) else None,
                            "fact_days": _float(fact_d_vals[i]) if i < len(fact_d_vals) else None,
                        })
                break
        return items

    # ==================================================================
    # CONTRACTORS (rows 22-29: col2=role, col22=company)
    # ==================================================================

    def _parse_contractors(self, ws: Worksheet) -> List[Dict]:
        contractors = []
        role_map = {
            "заказчик": "customer",
            "подрядчик по бурению": "drilling",
            "подрядчик по телеметри": "telemetry",
            "подрядчик по сопровождению раствор": "mud",
            "подрядчик по цементирован": "cementing",
            "подрядчик по долотам": "bits_gzd",
            "подрядчик по гти": "gti",
            "подрядчик по супервайзинг": "supervision",
        }

        for row in range(20, 35):
            label = _text(_cell(ws, row, 2))
            company = _text(_cell(ws, row, 22))
            if not label or not company:
                continue

            ll = label.lower()
            role = None
            for pattern, r in role_map.items():
                if pattern in ll:
                    role = r
                    break

            if role:
                contractors.append({"role": role, "company_name": company})

        # Evaluations at end of file
        for row in range(max(1, ws.max_row - 30), ws.max_row + 1):
            val = _text(_cell(ws, row, 2))
            if not val:
                continue
            vl = val.lower()

            for pattern, role in role_map.items():
                if pattern in vl or (role != "customer" and role.replace("_", " ") in vl):
                    eval_row = row + 1
                    eval_text = _text(_cell(ws, eval_row, 2))
                    if eval_text:
                        for c in contractors:
                            if c["role"] == role:
                                c["evaluation_text"] = eval_text
                    break

        return contractors

    # ==================================================================
    # DB operations
    # ==================================================================

    def _get_or_create_well(self, well_number: str, well_info: Dict, project_code: str) -> Well:
        well = self.db.query(Well).filter(Well.well_number == well_number).first()
        if not well:
            well = Well(
                well_number=well_number,
                well_name=f"{well_info.get('field_name', '')} / {well_number}".strip(" /"),
                field=well_info.get("field_name"),
                field_name=well_info.get("field_name"),
                project_code=project_code,
                company=well_info.get("customer", "ПАО Татнефть"),
            )
            self.db.add(well)
            self.db.flush()
        return well

    def _get_or_create_wellbore(self, well: Well, well_info: Dict) -> Wellbore:
        wellbore = self.db.query(Wellbore).filter(
            Wellbore.well_id == well.well_id,
            Wellbore.wellbore_number == "main",
        ).first()
        if not wellbore:
            wellbore = Wellbore(well_id=well.well_id, wellbore_number="main")
            self.db.add(wellbore)
            self.db.flush()
        return wellbore

    def _register_file(self, file_path: str, well: Well) -> Optional[File]:
        import os
        filename = os.path.basename(file_path)
        existing = self.db.query(File).filter(
            File.file_name == filename, File.well_id == well.well_id,
        ).first()
        if existing:
            return existing
        try:
            file_size = os.path.getsize(file_path)
        except OSError:
            file_size = None
        f = File(
            file_name=filename, file_path=file_path, file_type="xlsx",
            category="supervision_journal", well_id=well.well_id,
            file_size_bytes=file_size, processing_status="processing",
        )
        self.db.add(f)
        self.db.flush()
        return f

    def _update_well_passport(self, well: Well, info: Dict):
        simple_fields = [
            "well_purpose", "group_project", "target_horizon", "productive_layer",
            "altitude_rotor_m", "abs_mark_top_m", "design_depth_vertical_m",
            "design_depth_md_m", "magnetic_azimuth", "design_offset_m",
            "tolerance_radius_m", "rig_type", "mounting_type", "calendar_days",
            "unplanned_days", "npv_repair_days", "idle_days", "complication_days",
        ]
        for field in simple_fields:
            val = info.get(field)
            if val is not None:
                setattr(well, field, val)

        if info.get("field_name"):
            well.field = info["field_name"]
            well.field_name = info["field_name"]

        for date_field in ("drilling_start_date", "drilling_end_date"):
            d = info.get(date_field)
            if d:
                setattr(well, date_field, datetime.combine(d, time(0)) if isinstance(d, date) else d)

        if info.get("supervisors"):
            well.supervisors = info["supervisors"]

        well.has_supervision_log = True
        self.db.flush()

    def _update_wellbore_passport(self, wellbore: Wellbore, info: Dict):
        for field in ("abs_mark_bottom_m", "offset_to_bottom_m", "azimuth_to_bottom"):
            val = info.get(field)
            if val is not None:
                setattr(wellbore, field, val)
        v_vert = info.get("wellbore_design_depth_vertical_m")
        v_md = info.get("wellbore_design_depth_md_m")
        if v_vert:
            wellbore.design_depth_vertical_m = v_vert
        if v_md:
            wellbore.design_depth_md_m = v_md
        self.db.flush()
