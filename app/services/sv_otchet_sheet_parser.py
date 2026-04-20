"""
Парсер листа «Отчёт» (формат мультилистового журнала, в т.ч. final.xlsx).

Шапка «Суточный отчёт» и дата в R1 — шаблон; при импорте используйте
report_date (из Баланса/Детализации) или явный параметр.

Заполняет: sv_contractors, sv_rig_equipment, sv_well_construction (фрагменты),
sv_bha_runs, sv_drilling_regime, sv_mud_accounting, sv_chemical_reagents — все
привязаны к wellbore и/или report_id.
"""
from __future__ import annotations

import logging
import os
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from ..models import (
    Wellbore,
    SvDailyReport,
    SvContractor,
    SvRigEquipment,
    SvWellConstruction,
    SvBhaRun,
    SvDrillingRegime,
    SvMudAccounting,
    SvChemicalReagent,
    File,
)
from .sv_journal_parser import _float, _int, _text, _parse_date, _parse_range

logger = logging.getLogger(__name__)

SHEET_OTCHET = "Отчёт"

# Лист «Отчёт» в разных книгах может называться иначе
OTCHET_ALIASES = ("Отчёт", "Отчет", "Суточный отчет")


def _cell(ws: Worksheet, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def _find_sheet(wb: openpyxl.Workbook, name: Optional[str] = None) -> str:
    if name and name in wb.sheetnames:
        return name
    for n in OTCHET_ALIASES:
        if n in wb.sheetnames:
            return n
    for sn in wb.sheetnames:
        if "отч" in sn.lower():
            return sn
    raise ValueError("В книге не найден лист «Отчёт»")


def _parse_bha_components(desc: str) -> Tuple[Optional[str], Optional[float], Optional[str]]:
    bit_type = None
    bit_size = None
    motor_type = None
    if not desc:
        return bit_type, bit_size, motor_type
    bit_match = re.search(r"(\d{3}[.,]\d+)\s*PDC|PDC\s*[(\[]?([\d.,]+)", desc, re.IGNORECASE)
    if bit_match:
        size_str = bit_match.group(1) or bit_match.group(2)
        bit_size = _float(size_str)
        first_line = desc.split("\n")[0] if "\n" in desc else desc[:80]
        bit_type = first_line
    motor_match = re.search(r"(ВЗД[-\s]*\d+|ДР[-\s]*\d+[\s/\d]*)", desc, re.IGNORECASE)
    if motor_match:
        motor_type = motor_match.group(1).strip()
    return bit_type, bit_size, motor_type


def parse_header_date(ws: Worksheet) -> Optional[date]:
    """Дата в R1 (кол. 8) — справочно; может не совпадать с операционной датой."""
    v = _cell(ws, 1, 8)
    if isinstance(v, datetime):
        return v.date()
    return _parse_date(v)


def parse_contractors_gazprom(ws: Worksheet) -> List[Dict[str, str]]:
    """
    Строки 3–13: кол. 11 — роль, кол. 13 — организация / ФИО.
    """
    role_map = [
        ("заказчик", "customer"),
        ("представител", "customer_rep"),
        ("мастер буровой", "drilling_supervisor"),
        ("менеджер проекта", "supervision"),
        ("сервисные подрядчики", "service_contractors"),
        ("геофизик", "geophysics"),
        ("долотный сервис", "bits_service"),
        ("цгб", "mud_logging"),
    ]
    out: List[Dict[str, str]] = []
    for row in range(3, 20):
        label = _text(_cell(ws, row, 11))
        company = _text(_cell(ws, row, 13))
        if not label or not company:
            continue
        ll = label.lower()
        role = "other"
        for pat, r in role_map:
            if pat in ll:
                role = r
                break
        out.append({"role": role, "company_name": company[:2000]})
    return out


def parse_rig_equipment_gazprom(ws: Worksheet) -> Dict[str, Any]:
    equip: Dict[str, Any] = {}
    r3 = _text(_cell(ws, 3, 1))
    if r3 and "буровой станок" in r3.lower():
        v = _text(_cell(ws, 3, 4))
        if v:
            equip["rig_type"] = v
    return equip


def _first_range_in_row(ws: Worksheet, row: int) -> Tuple[Optional[float], Optional[float]]:
    """Ищет первое число или диапазон в колонках 4–12 (учёт объединённых ячеек)."""
    for col in range(4, 13):
        a, b = _parse_range(_cell(ws, row, col))
        if a is not None or b is not None:
            return a, b
    f = _float(_cell(ws, row, 8)) or _float(_cell(ws, row, 4))
    if f is not None:
        return f, f
    return None, None


def parse_drilling_regime_block(ws: Worksheet, max_row: int = 12) -> List[Dict[str, Any]]:
    """
    Блок «Режим бурения» — подписи в кол. 6, значения часто в 4–8 (в т.ч. после merge).
    """
    regimes: List[Dict[str, Any]] = []
    notes_parts: List[str] = []
    wob_min, wob_max = None, None
    p_min, p_max = None, None
    rpm_min, rpm_max = None, None
    flow = None

    for row in range(3, min(max_row + 1, ws.max_row + 1)):
        label = _text(_cell(ws, row, 6))
        if not label:
            continue
        ll = label.lower()
        unit = _text(_cell(ws, row, 10))
        ra, rb = _first_range_in_row(ws, row)
        raw_val = _cell(ws, row, 8)
        if raw_val is None and (ra is not None):
            raw_val = f"{ra}-{rb}" if rb is not None and ra != rb else ra

        if "нагрузка на долото" in ll:
            wob_min, wob_max = ra, rb if rb is not None else ra
            notes_parts.append(f"{label}: {raw_val} {unit or ''}".strip())
        elif "вес на крюке" in ll:
            notes_parts.append(f"{label}: {raw_val} {unit or ''}".strip())
        elif "давление" in ll and "вход" in ll:
            p_min, p_max = ra, rb if rb is not None else ra
            notes_parts.append(f"{label}: {raw_val} {unit or ''}".strip())
        elif "расход на входе" in ll:
            flow = ra if ra is not None else _float(raw_val)
            notes_parts.append(f"{label}: {raw_val} {unit or ''}".strip())
        elif "обороты" in ll:
            rpm_min, rpm_max = ra, rb if rb is not None else ra
            notes_parts.append(f"{label}: {raw_val} {unit or ''}".strip())
        elif "момент" in ll:
            notes_parts.append(f"{label}: {raw_val} {unit or ''}".strip())
        elif "мех" in ll and "скорость" in ll:
            notes_parts.append(f"{label}: {raw_val} {unit or ''}".strip())

    note = "; ".join(notes_parts) if notes_parts else None
    if any(x is not None for x in (wob_min, p_min, rpm_min, flow)):
        regimes.append({
            "status": "fact",
            "wob_min_ton": wob_min,
            "wob_max_ton": wob_max,
            "rpm_min": rpm_min,
            "rpm_max": rpm_max,
            "pressure_min": p_min,
            "pressure_max": p_max,
            "flow_rate_l_s": flow,
            "notes": note,
            "source_row": 3,
        })
    return regimes


def _find_row_contains(ws: Worksheet, col: int, text: str, start: int = 1, end: Optional[int] = None) -> Optional[int]:
    end = end or ws.max_row
    tl = text.lower()
    for r in range(start, min(end + 1, ws.max_row + 1)):
        v = _text(_cell(ws, r, col))
        if v and tl in v.lower():
            return r
    return None


def parse_bha_gazprom(ws: Worksheet) -> List[Dict[str, Any]]:
    """Таблица КНБК: от «Долото/Бур. головка» до «Параметры бурового раствора»."""
    start = _find_row_contains(ws, 1, "долото/бур", 1, 30)
    if not start:
        return []
    end = _find_row_contains(ws, 1, "параметры бурового раствора", start + 1)
    if not end:
        end = start + 120
    bha_list: List[Dict[str, Any]] = []
    current_status = "fact"
    for row in range(start + 1, min(end, ws.max_row + 1)):
        r1 = _text(_cell(ws, row, 1))
        if r1 and any(
            x in r1.lower()
            for x in ("№ долото", "№ спуска", "тип", "диаметр", "производитель", "интервал", "всего", "время бурения", "насадки", "код отработки")
        ):
            continue
        for col in range(1, 8):
            val = _text(_cell(ws, row, col))
            if val:
                vl = val.lower()
                if "план" in vl and "факт" not in vl:
                    current_status = "plan"
                elif "факт" in vl and "план" not in vl:
                    current_status = "fact"
        comp = _text(_cell(ws, row, 6))
        if not comp:
            parts = []
            for c in range(6, 17):
                t = _text(_cell(ws, row, c))
                if t and len(t) > 4:
                    parts.append(t)
            comp = "\n".join(parts) if parts else None
        if comp:
            cl = comp.lower()
            if any(x in cl for x in ("компоненты", "плановая кнбк", "обсадная колонна")):
                continue
        if comp and len(comp) > 8 and not comp.lower().startswith("плановая кнбк"):
            bt, bs, mt = _parse_bha_components(comp)
            num_val = _int(_cell(ws, row, 2))
            bha_list.append({
                "bha_number": num_val,
                "status": current_status,
                "description": comp[:12000],
                "bit_type": bt,
                "bit_size_mm": bs,
                "motor_type": mt,
                "source_row": row,
            })
    return bha_list


def parse_chemical_reagents_gazprom(ws: Worksheet) -> List[Dict[str, Any]]:
    """Колонки 17–22: наименование, ед., приход, расход за сутки, остаток, всего."""
    reagents: List[Dict[str, Any]] = []
    mud_start = _find_row_contains(ws, 1, "параметры бурового раствора", 1)
    limit_row = mud_start if mud_start else 200

    for row in range(5, min(limit_row, ws.max_row + 1)):
        name = _text(_cell(ws, row, 17))
        if not name or len(name) < 2:
            continue
        nl = name.lower()
        if "химреагент" in nl and "1." in nl:
            continue
        if any(k in nl for k in ("наименование", "итого", "всего")):
            continue
        unit = _text(_cell(ws, row, 18))
        rec = _float(_cell(ws, row, 19))
        used_day = _float(_cell(ws, row, 20))
        rem = _float(_cell(ws, row, 21))
        total_all = _float(_cell(ws, row, 22))

        if any(x is not None for x in (rec, used_day, rem, total_all)):
            reagents.append({
                "reagent_name": name[:500],
                "unit": unit,
                "total_received": rec,
                "used_preparation": None,
                "used_treatment": used_day,
                "used_regeneration": None,
                "exported": total_all,
                "remaining": rem,
                "source_row": row,
            })
    return reagents


def parse_mud_accounting_gazprom(ws: Worksheet) -> List[Dict[str, Any]]:
    """Блок «Буровой раствор» (строки ~55–61): учёт объёмов."""
    start = _find_row_contains(ws, 1, "буровой раствор", 40, 120)
    if not start:
        return []
    accounts: List[Dict[str, Any]] = []
    acc: Dict[str, Any] = {"mud_type": "Сводный учёт (Отчёт)", "source_row": start}

    for row in range(start, min(start + 15, ws.max_row + 1)):
        label = _text(_cell(ws, row, 1))
        if not label:
            continue
        ll = label.lower()
        v3 = _float(_cell(ws, row, 3))
        v4 = _float(_cell(ws, row, 4))
        v5 = _cell(ws, row, 5)
        s5 = _text(v5) if v5 is not None else None

        if "объём приготовленного" in ll:
            acc["volume_prepared"] = v4 if v4 is not None else v3
        elif "потери" in ll and "поверхности" in ll:
            acc["surface_losses"] = v4 if v4 is not None else v3
        elif "поглощение" in ll:
            acc["absorption"] = v4 if v4 is not None else v3
        elif "жфоб" in ll.replace(" ", ""):
            acc["total_losses"] = v4
        elif "тфоб" in ll.replace(" ", ""):
            acc["downhole_losses"] = v4
        elif "запас бур" in ll:
            acc["volume_remaining"] = _float(v5) if isinstance(v5, (int, float)) else _float(_text(v5))

    if len(acc) > 2:
        accounts.append(acc)
    return accounts


def parse_well_construction_snippets(ws: Worksheet) -> List[Dict[str, Any]]:
    """Обсадная колонка: строки с диаметром в кол. 12 и интервалом в кол. 13."""
    start = _find_row_contains(ws, 1, "долото/бур", 1, 30)
    if not start:
        return []
    end = _find_row_contains(ws, 1, "параметры бурового раствора", start + 1)
    if not end:
        end = start + 40
    items: List[Dict[str, Any]] = []
    for row in range(start + 2, min(end, ws.max_row + 1)):
        c12 = _text(_cell(ws, row, 12))
        c13 = _text(_cell(ws, row, 13))
        if not c12 or len(c12) < 6:
            continue
        if "диаметр" in c12.lower() and "марка" in c12.lower():
            continue
        if "*" in c12 or "мм" in c12.lower() or "OK" in c12 or "ОК" in c12:
            depth = None
            if c13:
                m = re.search(r"(\d+(?:[.,]\d+)?)\s*[-–/]\s*(\d+(?:[.,]\d+)?)", c13.replace(" ", ""))
                if m:
                    depth = _float(m.group(2))
            items.append({
                "casing_type": c12[:500],
                "outer_diameter_mm": None,
                "depth_m": depth,
            })
    return items[:50]


class SvOtchetSheetParserService:
    """Разбор листа «Отчёт» и запись в БД."""

    def __init__(self, db: Session):
        self.db = db
        self.warnings: List[str] = []

    def parse_preview(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sn = _find_sheet(wb, sheet_name)
        ws = wb[sn]

        contractors = parse_contractors_gazprom(ws)
        rig = parse_rig_equipment_gazprom(ws)
        regimes = parse_drilling_regime_block(ws)
        bha = parse_bha_gazprom(ws)
        reagents = parse_chemical_reagents_gazprom(ws)
        mud = parse_mud_accounting_gazprom(ws)
        construction = parse_well_construction_snippets(ws)
        hdr_date = parse_header_date(ws)

        return {
            "success": True,
            "sheet": sn,
            "header_date": hdr_date.isoformat() if hdr_date else None,
            "contractors_count": len(contractors),
            "rig_fields": list(rig.keys()),
            "drilling_regimes_count": len(regimes),
            "bha_rows_count": len(bha),
            "chemical_reagents_count": len(reagents),
            "mud_accounting_count": len(mud),
            "construction_items_count": len(construction),
            "warnings": self.warnings,
        }

    def import_sheet(
        self,
        file_path: str,
        wellbore_id: int,
        report_id: int,
        *,
        sheet_name: Optional[str] = None,
        replace_existing: bool = True,
        import_construction: bool = False,
    ) -> Dict[str, Any]:
        """
        Импорт в БД для уже существующего sv_daily_reports.report_id.

        replace_existing: удалить прежние дочерние строки по этому отчёту
        (КНБК, режим, раствор, реагенты) перед вставкой.
        import_construction: добавить строки sv_well_construction (по умолчанию выкл.,
        чтобы не дублировать при повторном импорте).
        """
        self.warnings = []
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sn = _find_sheet(wb, sheet_name)
        ws = wb[sn]

        rep = self.db.query(SvDailyReport).filter(SvDailyReport.report_id == report_id).first()
        if not rep or rep.wellbore_id != wellbore_id:
            raise ValueError("report_id не найден или не принадлежит wellbore_id")

        wb_row = self.db.query(Wellbore).filter(Wellbore.wellbore_id == wellbore_id).first()
        if not wb_row:
            raise ValueError("wellbore не найден")

        summary = {
            "report_id": report_id,
            "contractors_created": 0,
            "equipment_updated": 0,
            "construction_items_created": 0,
            "bha_runs_created": 0,
            "drilling_regimes_created": 0,
            "mud_accounting_created": 0,
            "chemical_reagents_created": 0,
            "warnings": self.warnings,
        }

        if replace_existing:
            for model, fk in (
                (SvBhaRun, "report_id"),
                (SvDrillingRegime, "report_id"),
                (SvMudAccounting, "report_id"),
                (SvChemicalReagent, "report_id"),
            ):
                self.db.query(model).filter(getattr(model, fk) == report_id).delete()
            self.db.flush()

        # Подрядчики (wellbore): upsert по role
        for c in parse_contractors_gazprom(ws):
            existing = self.db.query(SvContractor).filter(
                SvContractor.wellbore_id == wellbore_id,
                SvContractor.role == c["role"],
            ).first()
            if existing:
                existing.company_name = c["company_name"]
            else:
                self.db.add(SvContractor(wellbore_id=wellbore_id, **c))
            summary["contractors_created"] += 1

        equip = parse_rig_equipment_gazprom(ws)
        if equip:
            ex = self.db.query(SvRigEquipment).filter(
                SvRigEquipment.wellbore_id == wellbore_id,
            ).first()
            if ex:
                for k, v in equip.items():
                    setattr(ex, k, v)
            else:
                self.db.add(SvRigEquipment(wellbore_id=wellbore_id, **equip))
            summary["equipment_updated"] = 1

        if import_construction:
            for item in parse_well_construction_snippets(ws):
                self.db.add(SvWellConstruction(wellbore_id=wellbore_id, **item))
                summary["construction_items_created"] += 1

        for b in parse_bha_gazprom(ws):
            self.db.add(SvBhaRun(report_id=report_id, **b))
            summary["bha_runs_created"] += 1

        for r in parse_drilling_regime_block(ws):
            self.db.add(SvDrillingRegime(report_id=report_id, **r))
            summary["drilling_regimes_created"] += 1

        for m in parse_mud_accounting_gazprom(ws):
            self.db.add(SvMudAccounting(report_id=report_id, **m))
            summary["mud_accounting_created"] += 1

        for ch in parse_chemical_reagents_gazprom(ws):
            self.db.add(SvChemicalReagent(report_id=report_id, **ch))
            summary["chemical_reagents_created"] += 1

        self.db.commit()
        return summary


def register_otchet_file(db: Session, file_path: str, well_id: int) -> Optional[File]:
    """Опциональная регистрация файла (категория otchet_sheet)."""
    import os
    filename = os.path.basename(file_path)
    existing = db.query(File).filter(File.file_name == filename, File.well_id == well_id).first()
    if existing:
        return existing
    try:
        sz = os.path.getsize(file_path)
    except OSError:
        sz = None
    f = File(
        file_name=filename,
        file_path=file_path,
        file_type="xlsx",
        category="supervision_otchet_sheet",
        well_id=well_id,
        file_size_bytes=sz,
        processing_status="completed",
    )
    db.add(f)
    db.flush()
    return f
