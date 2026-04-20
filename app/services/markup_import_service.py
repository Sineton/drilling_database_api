"""
Service for importing markup xlsx files into normalized tables.
"""
from __future__ import annotations

import os
import re
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
from sqlalchemy.orm import Session

from ..models import (
    ActualOperation,
    Event,
    EventType,
    GeologyInterval,
    MarkupFileRow,
    Operation,
    Well,
    Wellbore,
)


WORKSHEET_NAME_FALLBACK = "Анализ данных"
ANNOTATION_SOURCE_DEFAULT = "markup_xlsx"

MARKUP_LAYOUT_EXTENDED = {
    "column_count": 34,
    "field_id": 1,
    "pad_id": 2,
    "source_well_number": 3,
    "lithology": 4,
    "formation_name": 5,
    "kg": 6,
    "hole_diameter_mm": 7,
    "bha_diameter_mm": 8,
    "inclination_deg": 9,
    "operation_label": 10,
    "risk_label": 14,
    "risk_level_id": 15,
    "markup_code": 16,
    "start_time": 17,
    "end_time": 18,
    "work_description": 19,
    "final_note": 33,
}

MARKUP_LAYOUT_COMPACT = {
    "column_count": 32,
    "field_id": None,
    "pad_id": None,
    "source_well_number": 1,
    "lithology": 2,
    "formation_name": 3,
    "kg": 4,
    "hole_diameter_mm": 5,
    "bha_diameter_mm": 6,
    "inclination_deg": 7,
    "operation_label": 8,
    "risk_label": 12,
    "risk_level_id": 13,
    "markup_code": 14,
    "start_time": 15,
    "end_time": 16,
    "work_description": 17,
    "final_note": 31,
}

OPERATION_CODE_DEFAULTS: Dict[str, Dict[str, Any]] = {
    "1.1.0": {"name": "Бурение (ротор)", "is_drilling": True, "risk_level": 0},
    "1.2.0": {"name": "Бурение (слайд)", "is_drilling": True, "risk_level": 0},
    "1.3.0": {"name": "Проработка", "is_drilling": True, "risk_level": 1},
    "1.4.0": {"name": "Наращивание", "is_drilling": False, "risk_level": 1},
    "1.5.0": {"name": "Промывка", "is_drilling": False, "risk_level": 1},
    "1.6.0": {"name": "Ориентирование ВЗД и ТС", "is_drilling": False, "risk_level": 1},
    "1.7.0": {"name": "Замер ТС", "is_drilling": False, "risk_level": 1},
    "1.8.0": {"name": "Спуск КНБК", "is_drilling": False, "risk_level": 2},
    "1.9.0": {"name": "Подъем КНБК", "is_drilling": False, "risk_level": 2},
    "1.10.0": {"name": "Спуск ОК", "is_drilling": False, "risk_level": 2},
    "1.11.0": {"name": "Шаблонировка", "is_drilling": False, "risk_level": 1},
    "1.12.0": {"name": "Разборка КНБК", "is_drilling": False, "risk_level": 1},
    "1.13.0": {"name": "Сборка КНБК", "is_drilling": False, "risk_level": 1},
}

EVENT_CODE_DEFAULTS: Dict[str, str] = {
    "2.1.1": "Затяжка до 5 т.",
    "2.1.2": "Затяжка более 5 т.",
    "2.2.1": "Посадка до 5 т.",
    "2.2.2": "Посадка более 5 т.",
    "2.3.1": "Отклонения параметров до X%",
    "2.3.2": "Отклонения параметров более X%",
    "2.4.1": "Отклонения параметров до X%",
    "2.4.2": "Отклонения параметров более X%",
    "2.5.1": "Отклонение параметра до 15%",
    "2.5.2": "Отклонение параметра более 15%",
    "2.6.1": "Простой до 5 мин.",
    "2.6.2": "Простой более 5 мин.",
    "3.1.0": "Прихват механический",
    "3.1.1": "Прихват: осыпи и обвалы",
    "3.1.2": "Прихват: запаковка",
    "3.1.3": "Сужение ствола / сложная геометрия / желобообразование",
    "3.1.4": "Прихват: пластичность пород",
    "3.1.5": "Прихват дифференциальный",
    "3.2.1": "Потеря циркуляции",
    "3.2.2": "Потеря циркуляции",
    "3.2.3": "Потеря циркуляции",
    "3.3.1": "Поглощение",
    "3.3.2": "Поглощение",
    "3.4": "ГНВП",
    "4.1": "Ликвидация осложнения",
}

INTERVAL_RE = re.compile(
    r"(?:инт\.?|интервал(?:е)?)\s*(?P<top>\d+(?:[.,]\d+)?)\s*[-–]\s*(?P<base>\d+(?:[.,]\d+)?)\s*м",
    re.IGNORECASE,
)


class MarkupImportService:
    """Parse and import markup workbook rows."""

    def __init__(self, db: Session):
        self.db = db

    def parse_preview(
        self,
        file_path: str,
        well_number_override: Optional[str] = None,
    ) -> Dict[str, Any]:
        parsed = self._parse_markup_workbook(file_path, well_number_override=well_number_override)
        resolved = self._resolve_codes(parsed["rows"])

        return {
            "sheet_name": parsed["sheet_name"],
            "well_number": parsed["well_number"],
            "total_rows": len(parsed["rows"]),
            "operation_rows": sum(1 for row in parsed["rows"] if row["operation_codes"]),
            "event_rows": sum(1 for row in parsed["rows"] if row["event_codes"]),
            "geology_candidates": sum(1 for row in parsed["rows"] if row["top_md"] is not None and row["base_md"] is not None),
            "unique_operation_codes": sorted(resolved["operation_codes"]),
            "unique_event_codes": sorted(resolved["event_codes"]),
            "missing_operation_codes": sorted(resolved["missing_operation_codes"]),
            "missing_event_codes": sorted(resolved["missing_event_codes"]),
            "samples": [self._sample_row(row) for row in parsed["rows"][:10]],
        }

    def import_markup(
        self,
        file_path: str,
        dry_run: bool = False,
        well_number_override: Optional[str] = None,
        project_code: Optional[str] = None,
        annotation_source: str = ANNOTATION_SOURCE_DEFAULT,
    ) -> Dict[str, Any]:
        parsed = self._parse_markup_workbook(file_path, well_number_override=well_number_override)
        well, wellbore = self._ensure_well_context(
            parsed["well_number"],
            project_code=project_code,
            dry_run=dry_run,
        )
        resolved = self._resolve_codes(parsed["rows"])

        missing_operation_codes = sorted(resolved["missing_operation_codes"])
        missing_event_codes = sorted(resolved["missing_event_codes"])

        warnings: List[str] = []
        errors: List[str] = []

        if missing_event_codes:
            warnings.append(
                "В файле найдены event codes, отсутствующие в БД: "
                + ", ".join(missing_event_codes)
            )

        operations_created = 0
        if missing_operation_codes and not dry_run:
            operations_created = self._ensure_operations(missing_operation_codes, parsed["rows"])

        event_types_created = 0
        if missing_event_codes and not dry_run:
            event_types_created = self._ensure_event_types(missing_event_codes, parsed["rows"])
            if event_types_created:
                missing_event_codes = []

        operation_map = {
            item.operation_code: item
            for item in self.db.query(Operation).filter(
                Operation.operation_code.in_(resolved["operation_codes"])
            ).all()
        } if resolved["operation_codes"] else {}

        event_type_map = {
            item.event_code: item
            for item in self.db.query(EventType).filter(
                EventType.event_code.in_(resolved["event_codes"])
            ).all()
        } if resolved["event_codes"] else {}

        if not dry_run:
            self.db.query(ActualOperation).filter(
                ActualOperation.wellbore_id == wellbore.wellbore_id,
                ActualOperation.source_file == os.path.basename(file_path),
            ).delete(synchronize_session=False)
            self.db.query(MarkupFileRow).filter(
                MarkupFileRow.wellbore_id == wellbore.wellbore_id,
                MarkupFileRow.source_file == os.path.basename(file_path),
            ).delete(synchronize_session=False)
            self.db.query(Event).filter(
                Event.wellbore_id == wellbore.wellbore_id,
                Event.annotation_source == annotation_source,
                Event.annotator_name == os.path.basename(file_path),
            ).delete(synchronize_session=False)

        geology_existing = self._existing_geology_keys(wellbore.wellbore_id)
        geology_created = 0
        summary_rows_created = 0
        actual_operations_created = 0
        events_created = 0

        for row in parsed["rows"]:
            operation_id = None
            if row["operation_codes"]:
                operation = operation_map.get(row["operation_codes"][0])
                if operation:
                    operation_id = operation.operation_id
                elif row["operation_codes"][0] not in missing_operation_codes:
                    warnings.append(
                        f"Не удалось сопоставить operation_code {row['operation_codes'][0]} "
                        f"для строки {row['row_number']}"
                    )

            event_type_ids: List[int] = []
            for code in row["event_codes"]:
                event_type = event_type_map.get(code)
                if event_type:
                    event_start_time = row["start_time"] or row["end_time"]
                    if event_start_time is None:
                        warnings.append(
                            f"Пропущено событие {code} в строке {row['row_number']}: отсутствует start_time/end_time"
                        )
                        continue

                    event_type_ids.append(event_type.event_type_id)
                    if not dry_run:
                        self.db.add(
                            Event(
                                wellbore_id=wellbore.wellbore_id,
                                event_type_id=event_type.event_type_id,
                                start_time=event_start_time,
                                end_time=row["end_time"],
                                start_md=row["top_md"],
                                end_md=row["base_md"],
                                annotation_source=annotation_source,
                                annotator_name=os.path.basename(file_path),
                                confidence=1.0,
                                notes=self._build_event_note(row),
                            )
                        )
                    events_created += 1
                else:
                    if code not in missing_event_codes:
                        missing_event_codes.append(code)

            geology_key = self._geology_key(
                wellbore_id=wellbore.wellbore_id,
                top_md=row["top_md"],
                base_md=row["base_md"],
                lithology=row["lithology"],
                formation_name=row["formation_name"],
                kg=row["kg"],
            )
            if geology_key and geology_key not in geology_existing:
                if not dry_run:
                    self.db.add(
                        GeologyInterval(
                            wellbore_id=wellbore.wellbore_id,
                            top_md=row["top_md"],
                            base_md=row["base_md"],
                            kg=row["kg"],
                            lithology=row["lithology"],
                            formation_name=row["formation_name"],
                        )
                    )
                geology_existing.add(geology_key)
                geology_created += 1

            if not dry_run:
                self.db.add(
                    ActualOperation(
                        wellbore_id=wellbore.wellbore_id,
                        operation_id=operation_id,
                        source_file=os.path.basename(file_path),
                        sequence_number=row["sequence_no"],
                        start_time=row["start_time"],
                        end_time=row["end_time"],
                        duration_minutes=self._duration_minutes(
                            row["start_time"],
                            row["end_time"],
                        ),
                        depth_from_m=row["top_md"],
                        depth_to_m=row["base_md"],
                        operation_code=row["operation_codes"][0] if row["operation_codes"] else None,
                        operation_label=row["operation_label"],
                        description=row["work_description"] or row["operation_label"] or "Операция из файла разметки",
                        risk_level_id=row["risk_level_id"],
                        markup_code=row["markup_code"],
                        event_codes=row["event_codes"] or None,
                    )
                )
                actual_operations_created += 1
                self.db.add(
                    MarkupFileRow(
                        source_file=os.path.basename(file_path),
                        wellbore_id=wellbore.wellbore_id,
                        source_well_number=row["source_well_number"],
                        sequence_no=row["sequence_no"],
                        field_id=row["field_id"],
                        pad_id=row["pad_id"],
                        lithology=row["lithology"],
                        formation_name=row["formation_name"],
                        kg=row["kg"],
                        hole_diameter_mm=row["hole_diameter_mm"],
                        bha_diameter_mm=row["bha_diameter_mm"],
                        inclination_deg=row["inclination_deg"],
                        operation_label=row["operation_label"],
                        risk_level_id=row["risk_level_id"],
                        markup_code=row["markup_code"],
                        operation_code=row["operation_codes"][0] if row["operation_codes"] else None,
                        operation_id=operation_id,
                        event_codes=row["event_codes"] or None,
                        event_type_ids=event_type_ids or None,
                        start_time=row["start_time"],
                        end_time=row["end_time"],
                        top_md=row["top_md"],
                        base_md=row["base_md"],
                        work_description=row["work_description"],
                        final_note=row["final_note"],
                    )
                )
                summary_rows_created += 1

        if dry_run:
            self.db.rollback()
        else:
            self.db.commit()

        if event_types_created:
            warnings.append(f"Созданы недостающие event types: {event_types_created}")
        if operations_created:
            warnings.append(f"Созданы недостающие operations: {operations_created}")

        return {
            "well_id": well.well_id,
            "wellbore_id": wellbore.wellbore_id,
            "well_number": parsed["well_number"],
            "dry_run": dry_run,
            "total_rows": len(parsed["rows"]),
            "operation_rows": sum(1 for row in parsed["rows"] if row["operation_codes"]),
            "event_rows": sum(1 for row in parsed["rows"] if row["event_codes"]),
            "operations_created": operations_created,
            "actual_operations_created": actual_operations_created if not dry_run else len(parsed["rows"]),
            "events_created": events_created,
            "geology_intervals_created": geology_created,
            "summary_rows_created": summary_rows_created if not dry_run else len(parsed["rows"]),
            "missing_operation_codes": sorted(set(missing_operation_codes)),
            "missing_event_codes": sorted(set(missing_event_codes)),
            "warnings": warnings,
            "errors": errors,
            "samples": [self._sample_row(row) for row in parsed["rows"][:10]],
        }

    def _parse_markup_workbook(
        self,
        file_path: str,
        well_number_override: Optional[str] = None,
    ) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet_name = self._get_markup_sheet_name(wb.sheetnames)
        ws = wb[sheet_name]
        layout = self._detect_layout(ws)

        rows: List[Dict[str, Any]] = []
        current_date: Optional[date] = None
        workbook_well_number: Optional[str] = None

        for row_idx, values in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            row = list(values[:layout["column_count"]])
            if len(row) < layout["column_count"]:
                row.extend([None] * (layout["column_count"] - len(row)))

            if not any(value is not None and str(value).strip() != "" for value in row):
                continue

            source_well_number = self._value_from_layout(row, layout, "source_well_number")
            if source_well_number and workbook_well_number is None:
                workbook_well_number = source_well_number

            start_time, current_date = self._normalize_datetime(
                self._raw_value_from_layout(row, layout, "start_time"),
                current_date,
            )
            end_time, current_date = self._normalize_datetime(
                self._raw_value_from_layout(row, layout, "end_time"),
                current_date,
            )
            if start_time and end_time and end_time < start_time:
                end_time += timedelta(days=1)

            work_description = self._value_from_layout(row, layout, "work_description")
            top_md, base_md = self._extract_interval_from_text(work_description)

            markup_code = self._value_from_layout(row, layout, "markup_code")
            code_parts = self._split_codes(markup_code)
            operation_codes = [code for code in code_parts if code.startswith("1.")]
            event_codes = [code for code in code_parts if not code.startswith("1.")]

            rows.append(
                {
                    "row_number": row_idx,
                    "sequence_no": self._to_int(row[0]),
                    "field_id": self._value_from_layout(row, layout, "field_id"),
                    "pad_id": self._value_from_layout(row, layout, "pad_id"),
                    "source_well_number": source_well_number,
                    "lithology": self._value_from_layout(row, layout, "lithology"),
                    "formation_name": self._value_from_layout(row, layout, "formation_name"),
                    "kg": self._to_float(self._raw_value_from_layout(row, layout, "kg")),
                    "hole_diameter_mm": self._to_float(self._raw_value_from_layout(row, layout, "hole_diameter_mm")),
                    "bha_diameter_mm": self._to_float(self._raw_value_from_layout(row, layout, "bha_diameter_mm")),
                    "inclination_deg": self._to_float(self._raw_value_from_layout(row, layout, "inclination_deg")),
                    "operation_label": self._value_from_layout(row, layout, "operation_label"),
                    "risk_label": self._value_from_layout(row, layout, "risk_label"),
                    "risk_level_id": self._to_int(self._raw_value_from_layout(row, layout, "risk_level_id")),
                    "markup_code": markup_code,
                    "operation_codes": operation_codes,
                    "event_codes": event_codes,
                    "start_time": start_time,
                    "end_time": end_time,
                    "top_md": top_md,
                    "base_md": base_md,
                    "work_description": work_description,
                    "final_note": self._value_from_layout(row, layout, "final_note"),
                }
            )

        well_number = str(well_number_override or workbook_well_number or "").strip()
        if not well_number:
            raise ValueError("Не удалось определить номер скважины из файла разметки")

        return {
            "sheet_name": sheet_name,
            "well_number": well_number,
            "rows": rows,
        }

    def _resolve_codes(self, rows: Iterable[Dict[str, Any]]) -> Dict[str, set]:
        operation_codes = sorted(
            {code for row in rows for code in row["operation_codes"] if code}
        )
        event_codes = sorted(
            {code for row in rows for code in row["event_codes"] if code}
        )

        operation_existing = {
            item[0]
            for item in self.db.query(Operation.operation_code).filter(
                Operation.operation_code.in_(operation_codes)
            ).all()
        } if operation_codes else set()

        event_existing = {
            item[0]
            for item in self.db.query(EventType.event_code).filter(
                EventType.event_code.in_(event_codes)
            ).all()
        } if event_codes else set()

        return {
            "operation_codes": set(operation_codes),
            "event_codes": set(event_codes),
            "missing_operation_codes": set(operation_codes) - operation_existing,
            "missing_event_codes": set(event_codes) - event_existing,
        }

    def _ensure_well_context(
        self,
        well_number: str,
        project_code: Optional[str],
        dry_run: bool,
    ) -> Tuple[Well, Wellbore]:
        well = self.db.query(Well).filter(Well.well_number == well_number).first()

        if well is None:
            if not project_code:
                raise ValueError(
                    f"Скважина {well_number} не найдена, а project_code для создания не передан"
                )
            well = Well(
                well_number=well_number,
                project_code=project_code,
                well_name=well_number,
                field_name=None,
            )
            if not dry_run:
                self.db.add(well)
                self.db.flush()
            else:
                well.well_id = 0

        wellbore = self.db.query(Wellbore).filter(
            Wellbore.well_id == well.well_id,
            Wellbore.wellbore_number == "main",
        ).first() if well.well_id else None

        if wellbore is None and well.well_id:
            wellbore = self.db.query(Wellbore).filter(
                Wellbore.well_id == well.well_id,
            ).order_by(Wellbore.wellbore_id.asc()).first()

            if wellbore is not None:
                wellbore.wellbore_number = "main"
                if not dry_run:
                    self.db.flush()

        if wellbore is None:
            wellbore = Wellbore(
                well_id=well.well_id if well.well_id else 0,
                wellbore_number="main",
            )
            if not dry_run:
                self.db.add(wellbore)
                self.db.flush()
            else:
                wellbore.wellbore_id = 0

        return well, wellbore

    def _ensure_operations(self, codes: List[str], rows: List[Dict[str, Any]]) -> int:
        created = 0
        names_by_code = {}
        for row in rows:
            for code in row["operation_codes"]:
                if code not in names_by_code:
                    names_by_code[code] = row["operation_label"]

        for code in codes:
            defaults = OPERATION_CODE_DEFAULTS.get(code, {})
            operation = Operation(
                operation_code=code,
                operation_name=defaults.get("name") or names_by_code.get(code) or code,
                is_drilling=defaults.get("is_drilling", False),
                risk_level=defaults.get("risk_level", 0),
                description="Автоматически создано при импорте файла разметки",
            )
            self.db.add(operation)
            created += 1

        self.db.flush()
        return created

    def _ensure_event_types(self, codes: List[str], rows: List[Dict[str, Any]]) -> int:
        created = 0
        labels_by_code = {}
        risks_by_code = {}

        for row in rows:
            for code in row["event_codes"]:
                labels_by_code.setdefault(code, row["operation_label"])
                risks_by_code.setdefault(code, row["risk_level_id"])

        for code in codes:
            category = self._event_category(code)
            event_type = EventType(
                event_code=code,
                event_name=EVENT_CODE_DEFAULTS.get(code) or labels_by_code.get(code) or code,
                parent_code=self._parent_code(code),
                category=category,
                is_complication=category in {"complication", "remediation"},
                is_precursor=category == "precursor",
                severity=risks_by_code.get(code) or 1,
                description="Автоматически создано при импорте файла разметки",
            )
            self.db.add(event_type)
            created += 1

        self.db.flush()
        return created

    def _existing_geology_keys(self, wellbore_id: int) -> set:
        items = self.db.query(GeologyInterval).filter(
            GeologyInterval.wellbore_id == wellbore_id
        ).all()
        return {
            self._geology_key(
                wellbore_id=item.wellbore_id,
                top_md=item.top_md,
                base_md=item.base_md,
                lithology=item.lithology,
                formation_name=item.formation_name,
                kg=item.kg,
            )
            for item in items
        }

    @staticmethod
    def _geology_key(
        wellbore_id: int,
        top_md: Optional[float],
        base_md: Optional[float],
        lithology: Optional[str],
        formation_name: Optional[str],
        kg: Optional[float],
    ) -> Optional[Tuple[Any, ...]]:
        if top_md is None or base_md is None:
            return None
        return (
            wellbore_id,
            round(min(top_md, base_md), 3),
            round(max(top_md, base_md), 3),
            (lithology or "").strip(),
            (formation_name or "").strip(),
            round(kg, 3) if kg is not None else None,
        )

    @staticmethod
    def _get_markup_sheet_name(sheet_names: List[str]) -> str:
        for name in sheet_names:
            if name == WORKSHEET_NAME_FALLBACK:
                return name
        for name in sheet_names:
            if "анализ" in name.lower():
                return name
        return sheet_names[0]

    @staticmethod
    def _detect_layout(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
        first_data_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True), None)
        row_len = len(first_data_row or ())
        if row_len >= MARKUP_LAYOUT_EXTENDED["column_count"]:
            return MARKUP_LAYOUT_EXTENDED
        return MARKUP_LAYOUT_COMPACT

    @staticmethod
    def _raw_value_from_layout(row: List[Any], layout: Dict[str, Any], field: str) -> Any:
        index = layout[field]
        if index is None:
            return None
        if index >= len(row):
            return None
        return row[index]

    @classmethod
    def _value_from_layout(
        cls,
        row: List[Any],
        layout: Dict[str, Any],
        field: str,
    ) -> Optional[str]:
        return cls._clean_str(cls._raw_value_from_layout(row, layout, field))

    @staticmethod
    def _split_codes(value: Optional[str]) -> List[str]:
        if not value:
            return []
        return [part.strip() for part in value.split(";") if part and part.strip()]

    @staticmethod
    def _normalize_datetime(
        value: Any,
        current_date: Optional[date],
    ) -> Tuple[Optional[datetime], Optional[date]]:
        if value is None:
            return None, current_date
        if isinstance(value, datetime):
            return value, value.date()
        if isinstance(value, date) and not isinstance(value, datetime):
            dt_value = datetime.combine(value, time.min)
            return dt_value, value
        if isinstance(value, time):
            if current_date is None:
                return None, None
            return datetime.combine(current_date, value), current_date
        if isinstance(value, str):
            clean = value.strip()
            if not clean:
                return None, current_date
            for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y %H:%M:%S"):
                try:
                    parsed = datetime.strptime(clean, fmt)
                    return parsed, parsed.date()
                except ValueError:
                    continue
        return None, current_date

    @staticmethod
    def _extract_interval_from_text(text: Optional[str]) -> Tuple[Optional[float], Optional[float]]:
        if not text:
            return None, None
        match = INTERVAL_RE.search(text)
        if not match:
            return None, None
        top = float(match.group("top").replace(",", "."))
        base = float(match.group("base").replace(",", "."))
        return min(top, base), max(top, base)

    @staticmethod
    def _build_event_note(row: Dict[str, Any]) -> str:
        pieces = [
            f"markup_code={row['markup_code']}" if row.get("markup_code") else None,
            f"risk_level_id={row['risk_level_id']}" if row.get("risk_level_id") is not None else None,
            row.get("work_description"),
            row.get("final_note"),
        ]
        return " | ".join(piece for piece in pieces if piece)

    @staticmethod
    def _duration_minutes(
        start_time: Optional[datetime],
        end_time: Optional[datetime],
    ) -> Optional[int]:
        if start_time is None or end_time is None:
            return None
        delta = end_time - start_time
        return max(int(delta.total_seconds() // 60), 0)

    @staticmethod
    def _sample_row(row: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "row_number": row["row_number"],
            "code": row["markup_code"],
            "operation_label": row["operation_label"],
            "risk_level_id": row["risk_level_id"],
            "start_time": row["start_time"],
            "end_time": row["end_time"],
            "top_md": row["top_md"],
            "base_md": row["base_md"],
            "description": row["work_description"],
        }

    @staticmethod
    def _event_category(code: str) -> str:
        if code.startswith("2."):
            return "precursor"
        if code.startswith("3."):
            return "complication"
        if code.startswith("4."):
            return "remediation"
        return "complication"

    @staticmethod
    def _parent_code(code: str) -> Optional[str]:
        parts = code.split(".")
        if len(parts) <= 1:
            return None
        parent = ".".join(parts[:-1]).strip(".")
        return parent or None

    @staticmethod
    def _clean_str(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _to_float(value: Any) -> Optional[float]:
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(",", "."))
        except ValueError:
            return None

    @staticmethod
    def _to_int(value: Any) -> Optional[int]:
        if value is None or value == "":
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        digits = str(value).strip()
        if not digits:
            return None
        try:
            return int(float(digits.replace(",", ".")))
        except ValueError:
            return None
