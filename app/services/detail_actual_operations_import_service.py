"""
Service for importing the "Детализация" sheet into actual_operations.
"""
from __future__ import annotations

import os
import re
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy.orm import Session

from ..models import ActualOperation, Operation, Well, Wellbore

SHEET_DETAIL = "Детализация"

HEADER_ALIASES = {
    "well_number": ("№ скв.", "№ скв"),
    "report_date": ("Дата",),
    "time_from": ("От",),
    "time_to": ("До",),
    "duration_hours": ("Время, час",),
    "description": ("Описание операции",),
    "section": ("Секция",),
    "stage": ("Этап",),
    "operation_label": ("Операция",),
    "npv_code": ("Код НПВ",),
    "incident_kind": ("Вид происшествия",),
    "incident_type": ("Тип происшествия",),
    "duration_days": ("Время, сут",),
    "pv_npv": ("ПВ/НПВ",),
}

INTERVAL_RE = re.compile(
    r"(?:инт\.?|интервал(?:е)?)\s*(?P<top>\d+(?:[.,]\d+)?)\s*[-–]\s*(?P<base>\d+(?:[.,]\d+)?)\s*м",
    re.IGNORECASE,
)


class DetailActualOperationsImportService:
    """Import actual operations from the "Детализация" worksheet."""

    def __init__(self, db: Session):
        self.db = db

    def import_sheet(
        self,
        file_path: str,
        dry_run: bool = False,
        well_number_override: Optional[str] = None,
        project_code: Optional[str] = None,
        replace_existing: bool = True,
    ) -> Dict[str, Any]:
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if SHEET_DETAIL not in workbook.sheetnames:
            raise ValueError(f"В файле нет листа '{SHEET_DETAIL}'")

        worksheet = workbook[SHEET_DETAIL]
        header_map = self._build_header_map(worksheet)
        parsed_rows, skipped_rows, warnings = self._parse_rows(worksheet, header_map)
        if not parsed_rows:
            raise ValueError("На листе 'Детализация' не найдено строк, пригодных для импорта")

        well_number = (well_number_override or self._resolve_well_number(parsed_rows) or "").strip()
        if not well_number:
            raise ValueError("Не удалось определить номер скважины. Укажите well_number.")

        well, wellbore = self._ensure_well_context(
            well_number=well_number,
            project_code=project_code,
            dry_run=dry_run,
        )

        source_file = os.path.basename(file_path)
        operation_map = self._operation_map()

        deleted_existing_rows = 0
        if not dry_run and replace_existing:
            deleted_existing_rows = (
                self.db.query(ActualOperation)
                .filter(
                    ActualOperation.wellbore_id == wellbore.wellbore_id,
                    ActualOperation.source_file == source_file,
                )
                .delete(synchronize_session=False)
            )

        imported_rows = 0
        matched_operations = 0
        unmatched_operations = 0
        rows_with_depth_interval = 0
        samples: List[Dict[str, Any]] = []

        for row in parsed_rows:
            matched_operation = operation_map.get(self._normalize_key(row["operation_label"]))
            if matched_operation:
                matched_operations += 1
            elif row["operation_label"]:
                unmatched_operations += 1

            if row["depth_from_m"] is not None and row["depth_to_m"] is not None:
                rows_with_depth_interval += 1

            if not dry_run:
                self.db.add(
                    ActualOperation(
                        wellbore_id=wellbore.wellbore_id,
                        operation_id=matched_operation.operation_id if matched_operation else None,
                        source_file=source_file,
                        sequence_number=row["sequence_number"],
                        start_time=row["start_time"],
                        end_time=row["end_time"],
                        duration_minutes=row["duration_minutes"],
                        depth_from_m=row["depth_from_m"],
                        depth_to_m=row["depth_to_m"],
                        operation_code=matched_operation.operation_code if matched_operation else None,
                        operation_label=row["operation_label"],
                        description=row["description"],
                        risk_level_id=None,
                        markup_code=None,
                        event_codes=None,
                    )
                )

            imported_rows += 1
            if len(samples) < 10:
                samples.append(
                    {
                        "sequence_number": row["sequence_number"],
                        "start_time": row["start_time"],
                        "end_time": row["end_time"],
                        "operation_label": row["operation_label"],
                        "description": row["description"],
                    }
                )

        if dry_run:
            self.db.rollback()
        else:
            self.db.commit()

        return {
            "well_id": well.well_id,
            "wellbore_id": wellbore.wellbore_id,
            "well_number": well_number,
            "dry_run": dry_run,
            "source_file": source_file,
            "total_rows": imported_rows + skipped_rows,
            "imported_rows": imported_rows,
            "skipped_rows": skipped_rows,
            "deleted_existing_rows": deleted_existing_rows,
            "matched_operations": matched_operations,
            "unmatched_operations": unmatched_operations,
            "rows_with_depth_interval": rows_with_depth_interval,
            "warnings": warnings,
            "samples": samples,
        }

    def _build_header_map(self, worksheet) -> Dict[str, int]:
        headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if headers is None:
            raise ValueError("Лист 'Детализация' пуст")

        normalized_headers = {
            self._normalize_header(header): index
            for index, header in enumerate(headers)
            if self._normalize_header(header)
        }

        header_map: Dict[str, int] = {}
        missing = []
        for field_name, aliases in HEADER_ALIASES.items():
            index = None
            for alias in aliases:
                index = normalized_headers.get(self._normalize_header(alias))
                if index is not None:
                    break
            if index is None and field_name in {"well_number", "report_date", "time_from", "time_to", "description", "operation_label"}:
                missing.append(aliases[0])
            header_map[field_name] = index if index is not None else -1

        if missing:
            raise ValueError("На листе 'Детализация' отсутствуют обязательные колонки: " + ", ".join(missing))
        return header_map

    def _parse_rows(self, worksheet, header_map: Dict[str, int]) -> Tuple[List[Dict[str, Any]], int, List[str]]:
        parsed_rows: List[Dict[str, Any]] = []
        warnings: List[str] = []
        skipped_rows = 0

        for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row = self._extract_row(values, header_map)
            if self._is_empty_row(row):
                skipped_rows += 1
                continue

            report_date = self._to_date(row["report_date_raw"])
            start_time = self._combine_date_and_time(report_date, row["time_from_raw"])
            end_time = self._combine_date_and_time(report_date, row["time_to_raw"])
            if start_time and end_time and end_time < start_time:
                end_time += timedelta(days=1)

            description = self._clean_text(row["description"])
            operation_label = self._clean_text(row["operation_label"])
            if not description or start_time is None or end_time is None:
                skipped_rows += 1
                warnings.append(
                    f"Строка {row_number} пропущена: требуется описание и корректные значения 'Дата'/'От'/'До'"
                )
                continue

            depth_from_m, depth_to_m = self._extract_interval_from_text(description)
            duration_minutes = self._duration_minutes(start_time, end_time)
            if duration_minutes is None:
                duration_minutes = self._hours_to_minutes(row["duration_hours_raw"])

            parsed_rows.append(
                {
                    "row_number": row_number,
                    "sequence_number": len(parsed_rows) + 1,
                    "well_number": self._clean_text(row["well_number"]),
                    "start_time": start_time,
                    "end_time": end_time,
                    "duration_minutes": duration_minutes,
                    "description": description,
                    "operation_label": operation_label,
                    "depth_from_m": depth_from_m,
                    "depth_to_m": depth_to_m,
                }
            )

        return parsed_rows, skipped_rows, warnings

    @staticmethod
    def _extract_row(values: Tuple[Any, ...], header_map: Dict[str, int]) -> Dict[str, Any]:
        def get_value(field_name: str) -> Any:
            index = header_map.get(field_name, -1)
            if index is None or index < 0 or index >= len(values):
                return None
            return values[index]

        return {
            "well_number": get_value("well_number"),
            "report_date_raw": get_value("report_date"),
            "time_from_raw": get_value("time_from"),
            "time_to_raw": get_value("time_to"),
            "duration_hours_raw": get_value("duration_hours"),
            "description": get_value("description"),
            "operation_label": get_value("operation_label"),
        }

    @staticmethod
    def _is_empty_row(row: Dict[str, Any]) -> bool:
        meaningful_values = [
            row.get("well_number"),
            row.get("report_date_raw"),
            row.get("time_from_raw"),
            row.get("time_to_raw"),
            row.get("description"),
            row.get("operation_label"),
        ]
        return not any(
            value is not None and str(value).strip() != ""
            for value in meaningful_values
        )

    def _resolve_well_number(self, rows: List[Dict[str, Any]]) -> Optional[str]:
        well_numbers = {
            row["well_number"]
            for row in rows
            if row.get("well_number")
        }
        if len(well_numbers) == 1:
            return next(iter(well_numbers))
        return None

    def _ensure_well_context(
        self,
        well_number: str,
        project_code: Optional[str],
        dry_run: bool,
    ) -> Tuple[Well, Wellbore]:
        well = self.db.query(Well).filter(Well.well_number == well_number).first()
        if well is None:
            if not project_code:
                raise ValueError(
                    f"Скважина {well_number} не найдена, а project_code для создания не передан"
                )
            well = Well(
                well_number=well_number,
                project_code=project_code,
                well_name=well_number,
            )
            if not dry_run:
                self.db.add(well)
                self.db.flush()
            else:
                well.well_id = 0

        wellbore = None
        if well.well_id:
            wellbore = self.db.query(Wellbore).filter(
                Wellbore.well_id == well.well_id,
                Wellbore.wellbore_number == "main",
            ).first()
            if wellbore is None:
                wellbore = self.db.query(Wellbore).filter(
                    Wellbore.well_id == well.well_id,
                ).order_by(Wellbore.wellbore_id.asc()).first()
                if wellbore is not None:
                    wellbore.wellbore_number = "main"
                    if not dry_run:
                        self.db.flush()

        if wellbore is None:
            wellbore = Wellbore(
                well_id=well.well_id if well.well_id else 0,
                wellbore_number="main",
            )
            if not dry_run:
                self.db.add(wellbore)
                self.db.flush()
            else:
                wellbore.wellbore_id = 0

        return well, wellbore

    def _operation_map(self) -> Dict[str, Operation]:
        operations = self.db.query(Operation).all()
        result: Dict[str, Operation] = {}
        for operation in operations:
            for key in (
                self._normalize_key(operation.operation_name),
                self._normalize_key(operation.operation_code),
            ):
                if key and key not in result:
                    result[key] = operation
        return result

    @staticmethod
    def _normalize_header(value: Any) -> str:
        text = str(value or "").strip().lower()
        return re.sub(r"\s+", " ", text)

    @staticmethod
    def _clean_text(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _normalize_key(value: Any) -> str:
        text = str(value or "").strip().lower().replace("ё", "е")
        return re.sub(r"\s+", " ", text)

    @staticmethod
    def _to_date(value: Any) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y", "%d.%m.%Y %H:%M", "%d.%m.%Y %H:%M:%S"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    @classmethod
    def _combine_date_and_time(cls, report_date: Optional[date], value: Any) -> Optional[datetime]:
        if value is None:
            return None
        if isinstance(value, datetime):
            if report_date is not None:
                return datetime.combine(report_date, value.time())
            return value
        if isinstance(value, time):
            if report_date is None:
                return None
            return datetime.combine(report_date, value)
        if isinstance(value, date):
            return datetime.combine(value, time.min)
        if isinstance(value, (int, float)) and report_date is not None and 0 <= float(value) < 1:
            seconds = int(round(float(value) * 24 * 60 * 60))
            time_value = (datetime.min + timedelta(seconds=seconds)).time()
            return datetime.combine(report_date, time_value)

        text = str(value).strip()
        if not text:
            return None

        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                if report_date is None:
                    return None
                parsed_time = datetime.strptime(text, fmt).time()
                return datetime.combine(report_date, parsed_time)
            except ValueError:
                continue

        for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"):
            try:
                parsed_dt = datetime.strptime(text, fmt)
                if report_date is not None:
                    return datetime.combine(report_date, parsed_dt.time())
                return parsed_dt
            except ValueError:
                continue

        return None

    @staticmethod
    def _hours_to_minutes(value: Any) -> Optional[int]:
        if value is None or value == "":
            return None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return max(int(round(float(value) * 60)), 0)
        try:
            return max(int(round(float(str(value).replace(",", ".")) * 60)), 0)
        except ValueError:
            return None

    @staticmethod
    def _duration_minutes(start_time: Optional[datetime], end_time: Optional[datetime]) -> Optional[int]:
        if start_time is None or end_time is None:
            return None
        return max(int((end_time - start_time).total_seconds() // 60), 0)

    @staticmethod
    def _extract_interval_from_text(text: Optional[str]) -> Tuple[Optional[float], Optional[float]]:
        if not text:
            return None, None
        match = INTERVAL_RE.search(text)
        if not match:
            return None, None
        top = float(match.group("top").replace(",", "."))
        base = float(match.group("base").replace(",", "."))
        return min(top, base), max(top, base)
